"""
ENERGY — Reporte Diario de Órdenes de Pedido
=============================================
Lee las órdenes desde ordenes.json en GitHub,
genera un informe HTML y lo envía por correo
vía Microsoft Graph API.

Ejecutado automáticamente por GitHub Actions
Lunes a Viernes a las 5:00 PM Colombia (UTC-5 = 22:00 UTC)
"""

import json
import os
import sys
import io
import base64
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


DAYS_ES   = {'Monday':'Lunes','Tuesday':'Martes','Wednesday':'Miércoles',
              'Thursday':'Jueves','Friday':'Viernes','Saturday':'Sábado','Sunday':'Domingo'}
MONTHS_ES = {'January':'enero','February':'febrero','March':'marzo','April':'abril',
              'May':'mayo','June':'junio','July':'julio','August':'agosto',
              'September':'septiembre','October':'octubre','November':'noviembre','December':'diciembre'}

GH_OWNER = 'cami902026-oss'
GH_REPO  = 'plataforma-eyg'

POC_STAGES = [
    {'key': 'compra',   'icon': '🛒', 'label': 'Compra'},
    {'key': 'entrega',  'icon': '🚚', 'label': 'Entrega'},
    {'key': 'cert',     'icon': '📋', 'label': 'Certificado'},
    {'key': 'factura',  'icon': '💰', 'label': 'Facturación'},
]


# ─── 1. LEER CREDENCIALES ─────────────────────────────────────────────────────

def load_config() -> dict:
    config_path = os.path.join(os.path.dirname(__file__), 'cowork_config.json')
    if not os.path.exists(config_path):
        return {}
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


# ─── 2. DESCARGAR ordenes.json DESDE GITHUB ───────────────────────────────────

def load_ordenes_from_github(gh_token: str) -> list:
    """Descarga ordenes.json del repositorio GitHub via API."""
    url = f'https://api.github.com/repos/{GH_OWNER}/{GH_REPO}/contents/ordenes.json'
    req = urllib.request.Request(url, headers={
        'Authorization': f'Bearer {gh_token}',
        'Accept': 'application/vnd.github+json',
        'User-Agent': 'EnergyBot/1.0'
    })
    try:
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
            content = base64.b64decode(data['content']).decode('utf-8')
            ordenes = json.loads(content)
            print(f'✅ ordenes.json descargado: {len(ordenes)} órdenes')
            return ordenes
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f'ERROR descargando ordenes.json: {e.code} — {body}')
        if e.code == 404:
            print('⚠️  ordenes.json no existe en el repo aún. Se enviará reporte vacío.')
            return []
        sys.exit(1)
    except Exception as ex:
        print(f'ERROR inesperado: {ex}')
        return []


# ─── 3. GENERAR EXCEL ────────────────────────────────────────────────────────

def generate_excel_op(ordenes: list) -> bytes | None:
    if not OPENPYXL_OK:
        print("⚠️  openpyxl no disponible — se omite el Excel adjunto")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Órdenes de Pedido'

    hdr_fill = PatternFill('solid', fgColor='0F2B5B')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    done_font = Font(color='1E7E34', bold=True)
    pend_font = Font(color='B7770D')

    headers = ['N° OP', 'Cliente', 'Descripción', 'Estado',
               'Compra', 'F. Compra', 'Entrega', 'F. Entrega',
               'Certificado', 'F. Cert.', 'Facturación', 'F. Factura']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, o in enumerate(ordenes, 2):
        stages = list(o.get('stages') or [])
        while len(stages) < 4:
            stages.append({})

        ws.cell(row=row_idx, column=1, value=o.get('num', '—'))
        ws.cell(row=row_idx, column=2, value=o.get('cliente', ''))
        ws.cell(row=row_idx, column=3, value=(o.get('desc') or '')[:120])
        ws.cell(row=row_idx, column=4, value=o.get('estado', 'activo').title())

        for si, st in enumerate(stages[:4]):
            col_base = 5 + si * 2
            done = is_stage_done(st)
            estado_txt = '✓ Hecho' if done else ('▶ Activo' if st.get('s') == 'active' else 'Pendiente')
            fecha_raw = st.get('f', '')
            fecha_txt = ''
            if fecha_raw:
                try:
                    p = fecha_raw.split('-')
                    fecha_txt = f"{p[2]}/{p[1]}/{p[0]}"
                except Exception:
                    fecha_txt = fecha_raw
            c_est = ws.cell(row=row_idx, column=col_base, value=estado_txt)
            c_est.font = done_font if done else pend_font
            ws.cell(row=row_idx, column=col_base + 1, value=fecha_txt)

        # Alternar fondo por fila
        if row_idx % 2 == 0:
            fill = PatternFill('solid', fgColor='EEF3FF')
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    col_widths = [12, 22, 42, 12, 12, 12, 12, 12, 14, 12, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'

    # Autofiltro en los títulos (permite filtrar/ordenar al abrir el Excel)
    last_col_letter = openpyxl.utils.get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── 4. GENERAR HTML DEL REPORTE ──────────────────────────────────────────────

def badge_estado(estado: str) -> str:
    colores = {
        'activo':     ('background:#e6f4ea;color:#1e7e34', 'Activo'),
        'completado': ('background:#e8f0fe;color:#1a56db', 'Completado'),
        'cancelado':  ('background:#fce8e6;color:#c0392b', 'Cancelado'),
    }
    style, label = colores.get(estado, ('background:#f0f0f0;color:#555', estado.title()))
    return f"<span style='padding:2px 10px;border-radius:12px;font-size:11px;font-weight:700;{style};'>{label}</span>"


def stage_dot(stage: dict, idx: int) -> str:
    s     = stage.get('s', 'pending')
    fecha = stage.get('f', '')
    nota  = stage.get('n', '')
    st    = POC_STAGES[idx]
    # done = s=='done' ó tiene fecha (igual que frontend)
    visually_done = (s == 'done' or bool(fecha.strip()))
    color = '#2EAA4A' if visually_done else ('#E8A020' if s == 'active' else '#8899bb')
    bg    = '#e6f4ea' if visually_done else ('#fff8e1' if s == 'active' else '#1e3a6e')
    fecha_txt = f"<br><span style='font-size:10px;color:#8899bb;'>{'/'.join(reversed(fecha.split('-')))}</span>" if fecha else ''
    nota_txt  = f"<br><span style='font-size:10px;color:#8899bb;' title='{nota}'>💬 {nota[:25]}{'…' if len(nota)>25 else ''}</span>" if nota else ''
    return (
        f"<td style='text-align:center;padding:4px 8px;'>"
        f"<div style='display:inline-block;background:{bg};border-radius:50%;width:28px;height:28px;"
        f"line-height:28px;font-size:14px;color:{color};border:2px solid {color};'>{st['icon']}</div>"
        f"<div style='font-size:10px;color:{color};font-weight:600;margin-top:2px;'>{st['label']}</div>"
        f"{fecha_txt}{nota_txt}</td>"
    )


def is_stage_done(st: dict) -> bool:
    """Una etapa está completa si s=='done' O si tiene fecha registrada (igual que el frontend)."""
    return st.get('s') == 'done' or bool(st.get('f', '').strip())


def get_etapa_actual(orden: dict) -> int:
    """Devuelve el índice (0-3) de la etapa actual de una orden activa.
    Busca la primera etapa 'active'; si no hay, la primera no-completada."""
    stages = orden.get('stages') or []
    # Asegurar 4 elementos
    while len(stages) < 4:
        stages.append({})
    # Primero: buscar etapa marcada 'active'
    for i, st in enumerate(stages[:4]):
        if st.get('s') == 'active':
            return i
    # Luego: primera etapa que NO esté done (done = s=='done' ó tiene fecha)
    for i, st in enumerate(stages[:4]):
        if not is_stage_done(st):
            return i
    return 3  # todas done → se queda en facturación


def clasificar_orden(orden: dict, today) -> tuple:
    """Clasifica una O.C. activa en su nivel de urgencia REAL.
    Retorna: (categoria, etapa_idx, fecha_obj_or_None)
    Categorías: 'vencida', 'sin_fecha', 'solo_cert', 'solo_factura', 'en_proceso'
    """
    stages = list(orden.get('stages') or [])
    while len(stages) < 4:
        stages.append({})
    pendientes = [i for i in range(4) if not is_stage_done(stages[i])]
    if not pendientes:
        return ('completa', None, None)

    # ¿Alguna etapa pendiente tiene fecha vencida?
    for i in pendientes:
        f = (stages[i].get('f') or '').strip()
        if f:
            try:
                fdt = datetime.strptime(f, '%Y-%m-%d').date()
                if fdt < today:
                    return ('vencida', i, fdt)
            except Exception:
                pass

    # Si solo certificado está pendiente → baja prioridad
    if pendientes == [2]:
        return ('solo_cert', 2, None)
    # Si solo facturación está pendiente → baja prioridad
    if pendientes == [3]:
        return ('solo_factura', 3, None)

    # Próxima etapa pendiente
    next_idx = pendientes[0]
    next_fecha = (stages[next_idx].get('f') or '').strip()
    if not next_fecha:
        return ('sin_fecha', next_idx, None)
    try:
        fdt = datetime.strptime(next_fecha, '%Y-%m-%d').date()
        return ('en_proceso', next_idx, fdt)
    except Exception:
        return ('sin_fecha', next_idx, None)


def build_resumen_etapas(activos: list) -> str:
    """Resumen accionable: clasifica cada O.C. activa en UNA categoría según urgencia real."""
    today = datetime.now().date()
    cats = {'vencida': [], 'sin_fecha': [], 'solo_cert': [], 'solo_factura': [], 'en_proceso': []}
    for o in activos:
        cat, idx, fecha = clasificar_orden(o, today)
        if cat in cats:
            num = o.get('num') or o.get('id', '—')
            extra = ''
            if cat == 'vencida' and fecha:
                dias = (today - fecha).days
                extra = f' (hace {dias}d)'
            elif cat == 'en_proceso' and fecha:
                dias = (fecha - today).days
                extra = f' (en {dias}d)'
            cats[cat].append((num, idx, extra))

    rows = [
        ('vencida',      '🚨', 'Vencidas',          '#e53e3e', 'Etapas con fecha pasada — requieren acción inmediata.'),
        ('sin_fecha',    '⚠️', 'Sin fecha',         '#E8A020', 'Próxima etapa sin fecha planificada.'),
        ('solo_cert',    '📋', 'Solo Certificado',  '#93c5fd', 'Únicamente falta el Certificado (baja prioridad — usualmente del cliente).'),
        ('solo_factura', '💰', 'Solo Facturación',  '#93c5fd', 'Únicamente falta la facturación.'),
        ('en_proceso',   '✅', 'En proceso al día', '#2EAA4A', 'Próxima etapa con fecha futura — todo en orden.'),
    ]
    filas_html = ''
    for key, icon, label, color, hint in rows:
        items = cats[key]
        cant = len(items)
        if cant == 0:
            lista = '—'
        else:
            etapa_nombres = ['Compra','Entrega','Certif.','Factura']
            partes = []
            for num, idx, extra in items[:20]:
                etapa = etapa_nombres[idx] if idx is not None else ''
                partes.append(f"<b>{num}</b>{f' · {etapa}' if etapa else ''}{extra}")
            lista = '<br>'.join(partes)
            if cant > 20:
                lista += f"<br><i>… y {cant-20} más</i>"
        filas_html += f"""
        <tr style='border-bottom:1px solid #1e3a6e;'>
          <td style='padding:12px 14px;font-size:22px;text-align:center;width:50px;'>{icon}</td>
          <td style='padding:12px 8px;'>
            <div style='font-size:13px;font-weight:700;color:{color};'>{label}</div>
            <div style='font-size:10px;color:#8899bb;margin-top:2px;'>{hint}</div>
          </td>
          <td style='padding:12px 8px;text-align:center;width:80px;'>
            <span style='font-size:24px;font-weight:900;color:{color};'>{cant}</span>
          </td>
          <td style='padding:12px 14px;font-size:11px;color:#cbd5ff;line-height:1.6;'>{lista}</td>
        </tr>"""

    return f"""
    <div style='margin-bottom:24px;'>
      <div style='font-size:11px;color:#8899bb;font-weight:700;text-transform:uppercase;
                  letter-spacing:1px;margin-bottom:10px;'>📊 Estado Real de las Órdenes Activas</div>
      <table style='width:100%;border-collapse:collapse;background:#0d1f3c;
                    border:1px solid #1e3a6e;border-radius:10px;overflow:hidden;'>
        <tbody>{filas_html}</tbody>
      </table>
    </div>"""


def build_report_html(ordenes: list, date_str: str) -> str:
    activos     = [o for o in ordenes if o.get('estado') == 'activo']
    completados = [o for o in ordenes if o.get('estado') == 'completado']
    cancelados  = [o for o in ordenes if o.get('estado') == 'cancelado']

    resumen_etapas = build_resumen_etapas(activos) if activos else ''

    resumen = f"""
    <div style='display:flex;gap:12px;flex-wrap:wrap;margin-bottom:24px;'>
      <div style='flex:1;min-width:120px;background:#0d1f3c;border:1px solid #1e3a6e;border-radius:10px;padding:16px;text-align:center;'>
        <div style='font-size:28px;font-weight:900;color:#fff;'>{len(ordenes)}</div>
        <div style='font-size:11px;color:#8899bb;text-transform:uppercase;letter-spacing:1px;margin-top:4px;'>Total</div>
      </div>
      <div style='flex:1;min-width:120px;background:#0d1f3c;border:1px solid #1e3a6e;border-radius:10px;padding:16px;text-align:center;'>
        <div style='font-size:28px;font-weight:900;color:#2EAA4A;'>{len(activos)}</div>
        <div style='font-size:11px;color:#8899bb;text-transform:uppercase;letter-spacing:1px;margin-top:4px;'>Activas</div>
      </div>
      <div style='flex:1;min-width:120px;background:#0d1f3c;border:1px solid #1e3a6e;border-radius:10px;padding:16px;text-align:center;'>
        <div style='font-size:28px;font-weight:900;color:#1a56db;'>{len(completados)}</div>
        <div style='font-size:11px;color:#8899bb;text-transform:uppercase;letter-spacing:1px;margin-top:4px;'>Completadas</div>
      </div>
      <div style='flex:1;min-width:120px;background:#0d1f3c;border:1px solid #1e3a6e;border-radius:10px;padding:16px;text-align:center;'>
        <div style='font-size:28px;font-weight:900;color:#e53e3e;'>{len(cancelados)}</div>
        <div style='font-size:11px;color:#8899bb;text-transform:uppercase;letter-spacing:1px;margin-top:4px;'>Canceladas</div>
      </div>
    </div>"""

    # Solo mostrar órdenes activas con al menos una etapa pendiente, ordenadas por urgencia
    today = datetime.now().date()
    PRIORIDAD = {'vencida':0, 'sin_fecha':1, 'solo_factura':2, 'solo_cert':3, 'en_proceso':4, 'completa':5}
    pendientes_clasif = []
    for o in activos:
        cat, idx, fecha = clasificar_orden(o, today)
        if cat == 'completa':
            continue
        pendientes_clasif.append((PRIORIDAD.get(cat, 9), cat, o))
    pendientes_clasif.sort(key=lambda x: x[0])
    pendientes = [p[2] for p in pendientes_clasif]

    if not pendientes:
        cuerpo = "<div style='text-align:center;padding:40px;color:#8899bb;'>✅ Todas las órdenes activas están al día.</div>"
    else:
        filas = ''
        for o in pendientes:
            stages = o.get('stages', [{},{},{},{}])
            while len(stages) < 4:
                stages.append({})
            dots = ''.join(stage_dot(stages[i], i) for i in range(4))
            filas += f"""
            <tr style='border-bottom:1px solid #1e3a6e;'>
              <td style='padding:12px 10px;'>
                <div style='font-weight:700;color:#fff;font-size:13px;'>{o.get('num','—')}</div>
                <div style='color:#8899bb;font-size:11px;margin-top:2px;'>{o.get('cliente','')}</div>
              </td>
              <td style='padding:12px 10px;color:#d0d9f0;font-size:12px;max-width:200px;'>
                {(o.get('desc') or '')[:80]}{'…' if len(o.get('desc',''))>80 else ''}
              </td>
              <td style='padding:12px 10px;text-align:center;'>{badge_estado(o.get('estado','activo'))}</td>
              {dots}
            </tr>"""

        cuerpo = f"""
        <table style='width:100%;border-collapse:collapse;background:#0d1f3c;border-radius:10px;overflow:hidden;'>
          <thead>
            <tr style='background:#0F2B5B;'>
              <th style='padding:10px;text-align:left;font-size:11px;color:#E8A020;text-transform:uppercase;letter-spacing:1px;' colspan='7'>⚠️ Órdenes con Etapas Pendientes</th></tr><tr style='background:#0F2B5B;'>
              <th style='padding:10px;text-align:left;font-size:11px;color:#8899bb;text-transform:uppercase;letter-spacing:1px;'>OP / Proyecto</th>
              <th style='padding:10px;text-align:left;font-size:11px;color:#8899bb;text-transform:uppercase;letter-spacing:1px;'>Descripción</th>
              <th style='padding:10px;text-align:center;font-size:11px;color:#8899bb;text-transform:uppercase;letter-spacing:1px;'>Estado</th>
              <th style='padding:10px;text-align:center;font-size:11px;color:#8899bb;text-transform:uppercase;letter-spacing:1px;'>🛒 Compra</th>
              <th style='padding:10px;text-align:center;font-size:11px;color:#8899bb;text-transform:uppercase;letter-spacing:1px;'>🚚 Entrega</th>
              <th style='padding:10px;text-align:center;font-size:11px;color:#8899bb;text-transform:uppercase;letter-spacing:1px;'>📋 Cert.</th>
              <th style='padding:10px;text-align:center;font-size:11px;color:#8899bb;text-transform:uppercase;letter-spacing:1px;'>💰 Factura</th>
            </tr>
          </thead>
          <tbody>{filas}</tbody>
        </table>"""

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
</head>
<body style="margin:0;padding:0;background:#f4f6fb;font-family:'Open Sans',Arial,sans-serif;">
<div style="max-width:900px;margin:0 auto;padding:16px;">
  <div style="background:#0F2B5B;padding:18px 28px;border-radius:12px 12px 0 0;display:flex;justify-content:space-between;align-items:center;">
    <div>
      <div style="font-size:20px;font-weight:900;color:#fff;letter-spacing:1px;">⚡ ENERGY</div>
      <div style="font-size:11px;color:#93c5fd;margin-top:2px;text-transform:uppercase;letter-spacing:1px;">Reporte Diario — Órdenes de Pedido</div>
    </div>
    <div style="font-size:12px;color:#93c5fd;text-align:right;">{date_str}</div>
  </div>
  <div style="height:3px;background:linear-gradient(90deg,#E8A020,transparent);"></div>
  <div style="background:#0a1a30;padding:12px 20px;text-align:center;border-bottom:1px solid #1e3a6e;">
    <a href="https://cami902026-oss.github.io/plataforma-eyg/Index.html" target="_blank"
       style="display:inline-block;background:#2EAA4A;color:#fff;padding:10px 28px;border-radius:8px;font-weight:700;font-size:14px;text-decoration:none;letter-spacing:0.5px;">
      🔍 Ver informe interactivo con búsqueda
    </a>
    <div style="font-size:11px;color:#8899bb;margin-top:6px;">Abre la plataforma en tu navegador para buscar y filtrar órdenes</div>
  </div>
  <div style="background:#071525;padding:20px;border-radius:0 0 0 0;">
    {resumen}
    {resumen_etapas}
    {cuerpo}
  </div>
  <div style="background:#071525;padding:14px;border-radius:0 0 12px 12px;text-align:center;margin-top:0;border-top:1px solid #1e3a6e;">
    <p style="margin:3px 0;font-size:11px;color:#8899bb;">⚡ Generado por <strong style="color:#E8A020;">ENERGY — Asistente Administrativo</strong></p>
    <p style="margin:3px 0;font-size:11px;color:#8899bb;">E&amp;G Energy Group · Reporte automático L–V 5:00 PM Colombia</p>
  </div>
</div>
</body>
</html>"""


# ─── 4. AUTENTICACIÓN MICROSOFT GRAPH ─────────────────────────────────────────

def get_access_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    url = f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token'
    data = urllib.parse.urlencode({
        'grant_type':    'client_credentials',
        'client_id':     client_id,
        'client_secret': client_secret,
        'scope':         'https://graph.microsoft.com/.default'
    }).encode()
    req = urllib.request.Request(url, data=data, method='POST')
    print("🔑 Obteniendo token Microsoft Graph...")
    try:
        with urllib.request.urlopen(req) as resp:
            print("✅ Token obtenido correctamente")
            return json.loads(resp.read())['access_token']
    except urllib.error.HTTPError as e:
        print(f"ERROR obteniendo token: {e.code} — {e.read().decode()}")
        sys.exit(1)


# ─── 5. ENVIAR CORREO VIA GRAPH ────────────────────────────────────────────────

def send_email(token: str, sender: str, recipients: list, subject: str, html_body: str,
               attachment_bytes: bytes = None, attachment_name: str = None):
    msg = {
        'subject': subject,
        'body': {'contentType': 'HTML', 'content': html_body},
        'toRecipients': [{'emailAddress': {'address': r.strip()}} for r in recipients]
    }
    if attachment_bytes and attachment_name:
        msg['attachments'] = [{
            '@odata.type': '#microsoft.graph.fileAttachment',
            'name': attachment_name,
            'contentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'contentBytes': base64.b64encode(attachment_bytes).decode('utf-8')
        }]
    payload = json.dumps({'message': msg, 'saveToSentItems': True}).encode('utf-8')
    url = f'https://graph.microsoft.com/v1.0/users/{sender}/sendMail'
    req = urllib.request.Request(url, data=payload, method='POST',
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'})
    try:
        with urllib.request.urlopen(req) as resp:
            print(f"✅ Correo enviado (HTTP {resp.status})")
    except urllib.error.HTTPError as e:
        print(f"ERROR enviando correo: {e.code} — {e.read().decode()}")
        sys.exit(1)


# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    cfg = load_config()

    tenant_id     = os.environ.get('MS_TENANT_ID',     cfg.get('ms_tenant_id',     '')).strip()
    client_id     = os.environ.get('MS_CLIENT_ID',     cfg.get('ms_client_id',     '')).strip()
    client_secret = os.environ.get('MS_CLIENT_SECRET', cfg.get('ms_client_secret', '')).strip()
    sender_email  = os.environ.get('SENDER_EMAIL',     cfg.get('sender_email',     '')).strip()
    recipients    = [r.strip() for r in os.environ.get(
                        'RECIPIENT_EMAILS', cfg.get('recipient_emails', '')).split(',')]
    extra_str     = os.environ.get('EXTRA_RECIPIENTS', cfg.get('extra_recipients', '')).strip()
    if extra_str:
        recipients += [r.strip() for r in extra_str.split(',') if r.strip()]
    recipients    = list(dict.fromkeys(r for r in recipients if r))  # eliminar duplicados vacíos
    gh_token      = os.environ.get('GH_TOKEN',         cfg.get('github_token',     '')).strip()

    print(f"🔍 Sender        : {sender_email}")
    print(f"🔍 Destinatarios : {', '.join(recipients)}")

    # Descargar órdenes desde GitHub
    ordenes = load_ordenes_from_github(gh_token)

    now      = datetime.now()
    day      = DAYS_ES.get(now.strftime('%A'), now.strftime('%A'))
    month    = MONTHS_ES.get(now.strftime('%B'), now.strftime('%B'))
    date_str = f"{day} {now.day} de {month} de {now.year}"
    subject  = f"📋 Informe OP Activas E&G — {date_str}"

    print(f"📝 Generando informe para {len(ordenes)} órdenes...")
    email_html = build_report_html(ordenes, date_str)

    excel_bytes = generate_excel_op(ordenes)
    excel_name  = f"OP_EyG_{now.strftime('%Y-%m-%d')}.xlsx"
    if excel_bytes:
        print(f"📊 Excel generado: {excel_name} ({len(excel_bytes)//1024} KB)")
    else:
        print("⚠️  Excel no generado (openpyxl no disponible)")

    token = get_access_token(tenant_id, client_id, client_secret)
    print(f"📧 Enviando a: {', '.join(recipients)}")
    send_email(token, sender_email, recipients, subject, email_html, excel_bytes, excel_name)
    print("🎉 Informe OP enviado exitosamente.")
