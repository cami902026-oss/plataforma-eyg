# -*- coding: utf-8 -*-
"""
INFORME COMERCIAL SEMANAL — E&G Energy Group
=============================================
Se envía los VIERNES 6:00 PM (hora Colombia).

Para  : Gerencia
Copia : Andrea + cada vendedor

Contenido (solo cotizaciones — sin O.C. ni inventario, a propósito):
  1. Resumen de la semana (lunes a viernes)
  2. Las cotizaciones de la semana, agrupadas por vendedor, con su estado
  3. Pendientes de cerrar: vencidas ACUMULADAS DEL MES que siguen sin estado final
  4. Adjudicadas que todavía no se facturan
  5. Tabla resumen por vendedor
  + Excel adjunto con las tres listas completas

Es informativo. No escribe nada: solo LEE de Supabase.

Prueba local sin enviar correo:
    TEST_OUT=%TEMP% PYTHONIOENCODING=utf-8 python scripts/informe_semanal_comercial.py
"""

import os
import json
import base64
import datetime
import urllib.request
import urllib.parse

SUPABASE_URL_DEFAULT = 'https://juprjevxkcitqpsnemto.supabase.co'
SUPABASE_KEY_DEFAULT = 'sb_publishable_zZrmpmvqbz4AJCGHRHQ8Xw_8tnf5ObM'

# Una cotización sin fecha de vencimiento se considera vencida a los 30 días del envío.
# Mismo criterio que la bandeja "⏰ Sin cerrar" de la plataforma, para que los números cuadren.
DIAS_FALLBACK = 30
ESTADOS_ABIERTOS = ('Enviada', 'Revisión gerencia', 'Pendiente')

MESES = ['ene', 'feb', 'mar', 'abr', 'may', 'jun', 'jul', 'ago', 'sep', 'oct', 'nov', 'dic']

# Colores corporativos E&G (los mismos de la plataforma)
AZUL = '#1A3A8F'
ORO = '#E8A020'
VERDE = '#1F7A38'
ROJO = '#A32B1E'
GRIS = '#6B7A90'
LINEA = '#D5DCE6'


# ─── UTILIDADES ───────────────────────────────────────────────────────────────

def now_co():
    return datetime.datetime.now(datetime.timezone.utc).replace(tzinfo=None) - datetime.timedelta(hours=5)


def money(n):
    try:
        return '$' + format(int(round(float(n or 0))), ',d').replace(',', '.')
    except Exception:
        return '$0'


def fecha_larga(d):
    return '%d de %s' % (d.day, ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio',
                                 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre'][d.month - 1])


def fecha_corta(d):
    return '%d-%s' % (d.day, MESES[d.month - 1])


def parse_fecha(s):
    """'2026-08-22' o '2026-08-22T...' -> date. None si no se puede."""
    s = (s or '')[:10]
    try:
        y, m, d = s.split('-')
        return datetime.date(int(y), int(m), int(d))
    except Exception:
        return None


def fetch_supabase_paged(table, params=''):
    base = os.environ.get('SUPABASE_URL', SUPABASE_URL_DEFAULT).strip().rstrip('/')
    key = os.environ.get('SUPABASE_KEY', SUPABASE_KEY_DEFAULT).strip()
    out, offset, page = [], 0, 1000
    while True:
        url = '%s/rest/v1/%s?limit=%d&offset=%d%s' % (base, table, page, offset, params)
        req = urllib.request.Request(url, headers={'apikey': key, 'Authorization': 'Bearer ' + key})
        try:
            with urllib.request.urlopen(req) as resp:
                chunk = json.loads(resp.read())
        except Exception as e:
            print('⚠️  Supabase %s (offset %d): %s' % (table, offset, e))
            break
        if not isinstance(chunk, list) or not chunk:
            break
        out.extend(chunk)
        if len(chunk) < page:
            break
        offset += page
    print('   Supabase %s: %d filas' % (table, len(out)))
    return out


# ─── LECTURA Y NORMALIZACIÓN ──────────────────────────────────────────────────

def cargar_cotizaciones():
    """Trae la cotización COMPLETA desde la columna `extra` (jsonb), que es el
    registro tal como lo guarda la plataforma. Si una fila vieja no tiene `extra`,
    se reconstruye lo mínimo desde las columnas sueltas."""
    filas = fetch_supabase_paged('cotizaciones', '&deleted=is.false&select=id,fecha,estado,cliente,'
                                 'vendedor,realizada_por,total,subtotal,iva,fecha_venc,fecha_envio,'
                                 'motivo_rechazo,adjudicada_at,facturada_at,valor_adjudicado,factura,'
                                 'fuente,extra')
    cots = []
    for f in filas:
        ex = f.get('extra') if isinstance(f.get('extra'), dict) else {}
        c = {
            'id': f.get('id') or ex.get('id') or '',
            'fecha': (f.get('fecha') or ex.get('fecha') or '')[:10],
            'estado': f.get('estado') or ex.get('estado') or '',
            'cliente': (f.get('cliente') or ex.get('cliente') or '').strip(),
            'vendedor': (f.get('vendedor') or ex.get('vendedor') or '').strip(),
            'realizadaPor': (f.get('realizada_por') or ex.get('realizadaPor') or '').strip(),
            'total': f.get('total') if f.get('total') is not None else ex.get('total'),
            'subtotal': f.get('subtotal') if f.get('subtotal') is not None else ex.get('subtotal'),
            'iva': f.get('iva') if f.get('iva') is not None else ex.get('iva'),
            'fechaVenc': (f.get('fecha_venc') or ex.get('fechaVenc') or '')[:10],
            'fechaEnvio': (f.get('fecha_envio') or ex.get('fechaEnvio') or '')[:10],
            'motivoRechazo': f.get('motivo_rechazo') or ex.get('motivoRechazo') or '',
            'adjudicadaAt': (f.get('adjudicada_at') or ex.get('adjudicadaAt') or '')[:10],
            'facturadaAt': (f.get('facturada_at') or ex.get('facturadaAt') or '')[:10],
            'valorAdjudicado': f.get('valor_adjudicado') if f.get('valor_adjudicado') is not None else ex.get('valorAdjudicado'),
            'factura': (f.get('factura') or ex.get('factura') or '').strip(),
            'fuente': f.get('fuente') or '',
        }
        cots.append(c)
    return cots


def base_sin_iva(c):
    """Base real de negocio. El valor adjudicado que guarda la plataforma es
    suma(cantidad x precio) = SIN IVA, así que todo se compara contra el subtotal."""
    try:
        s = float(c.get('subtotal') or 0)
        if s > 0:
            return s
    except Exception:
        pass
    t = float(c.get('total') or 0)
    try:
        iva = float(c.get('iva') or 0)
    except Exception:
        iva = 0
    return t / (1 + iva / 100) if iva > 0 else t


def vendedor_de(c):
    return c.get('vendedor') or c.get('realizadaPor') or '⚠️ Sin vendedor'


def dias_vencida(c, hoy):
    """Días que lleva vencida, o None si aún está vigente / no se puede calcular."""
    fv = parse_fecha(c.get('fechaVenc'))
    base = parse_fecha(c.get('fechaEnvio')) or parse_fecha(c.get('fecha'))
    # Vencimientos basura (p.ej. 1900-01-08, seriales de Excel mal convertidos):
    # si la cotización vence antes de existir, se ignora y se usa la regla de 30 días.
    if fv and (base is None or fv >= base):
        limite = fv
    elif base:
        limite = base + datetime.timedelta(days=DIAS_FALLBACK)
    else:
        return None
    d = (hoy - limite).days
    return d if d > 0 else None


def valor_ganado(c):
    """Lo realmente adjudicado si se registró; si no, el cotizado (las viejas)."""
    v = c.get('valorAdjudicado')
    if v is not None:
        try:
            return float(v)
        except Exception:
            pass
    return base_sin_iva(c)


# ─── HTML ─────────────────────────────────────────────────────────────────────
# Correo: todo en tablas con estilos en línea. Outlook no soporta flexbox ni
# variables CSS, así que nada de eso aquí.

CHIP = {
    'Enviada':           ('#E7ECF7', AZUL,  'Enviada'),
    'Adjudicada':        ('#E3F3E7', VERDE, 'Adjudicada'),
    'Facturada':         ('#D4E9DC', '#15602B', 'Facturada'),
    'Aprobada':          ('#E3F3E7', VERDE, 'Aprobada'),
    'Rechazada':         ('#FBE7E4', ROJO,  'Rechazada'),
    'Vencida':           ('#FBF0D9', '#9A6A05', 'Vencida'),
    'Anulada':           ('#FBE7E4', ROJO,  'Anulada'),
    'Borrador':          ('#EEF1F5', GRIS,  'Borrador'),
    'Revisión gerencia': ('#EFE7FB', '#5B3A9B', 'Rev. gerencia'),
    'Pendiente':         ('#EEF1F5', GRIS,  'Pendiente'),
}


def esc(s):
    return (str(s or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))


def chip(estado):
    bg, fg, txt = CHIP.get(estado, ('#EEF1F5', GRIS, estado or '—'))
    return ('<span style="display:inline-block;background:%s;color:%s;font-size:11px;'
            'font-weight:bold;padding:2px 8px;border-radius:3px;white-space:nowrap;">%s</span>'
            % (bg, fg, esc(txt)))


def card(num, label, color):
    return ('<td style="padding:0 6px 12px 0;" valign="top" width="20%%">'
            '<table cellpadding="0" cellspacing="0" border="0" width="100%%" '
            'style="background:#ffffff;border:1px solid %s;border-left:3px solid %s;border-radius:3px;">'
            '<tr><td style="padding:12px 14px;">'
            '<div style="font-size:10px;color:%s;text-transform:uppercase;letter-spacing:1px;'
            'font-weight:bold;padding-bottom:5px;">%s</div>'
            '<div style="font-size:21px;color:#0F1B2D;font-weight:bold;line-height:1.1;">%s</div>'
            '</td></tr></table></td>' % (LINEA, color, GRIS, esc(label), esc(num)))


def th(txt, align='left'):
    return ('<th align="%s" style="font-size:10px;color:%s;text-transform:uppercase;'
            'letter-spacing:.8px;padding:8px 10px;border-bottom:1px solid %s;'
            'background:#F6F8FB;white-space:nowrap;">%s</th>' % (align, GRIS, LINEA, esc(txt)))


def td(txt, align='left', extra=''):
    return ('<td align="%s" style="font-size:13px;color:#0F1B2D;padding:8px 10px;'
            'border-bottom:1px solid #E9EDF3;%s">%s</td>' % (align, extra, txt))


def titulo(txt, sub=''):
    s = ('<div style="font-size:12.5px;color:%s;padding-top:3px;">%s</div>' % (GRIS, esc(sub))) if sub else ''
    return ('<tr><td style="padding:26px 0 10px 0;">'
            '<div style="font-size:17px;color:#0F1B2D;font-weight:bold;border-top:2px solid %s;'
            'padding-top:12px;">%s</div>%s</td></tr>' % (AZUL, esc(txt), s))


def build_html(d):
    lunes, viernes = d['lunes'], d['viernes']
    H = []
    A = H.append

    A('<div style="background:#EEF1F5;padding:0;margin:0;">')
    A('<table cellpadding="0" cellspacing="0" border="0" width="100%" style="background:#EEF1F5;'
      'font-family:Arial,Helvetica,sans-serif;">')
    A('<tr><td align="center" style="padding:20px 12px;">')
    A('<table cellpadding="0" cellspacing="0" border="0" width="820" style="max-width:820px;width:100%;">')

    # ── Encabezado ──
    A('<tr><td style="background:%s;padding:22px 24px;border-radius:3px 3px 0 0;">' % AZUL)
    A('<div style="font-size:11px;color:#B9CBF0;text-transform:uppercase;letter-spacing:2px;'
      'font-weight:bold;">E&amp;G Energy Group &middot; Informe semanal</div>')
    A('<div style="font-size:25px;color:#ffffff;font-weight:bold;padding-top:6px;line-height:1.2;">'
      'Cotizaciones de la semana</div>')
    A('<div style="font-size:13px;color:#D6E1F7;padding-top:6px;">%s al %s de %d</div>'
      % (fecha_larga(lunes), fecha_larga(viernes), viernes.year))
    A('</td></tr>')
    A('<tr><td style="background:#ffffff;padding:22px 24px;border:1px solid %s;border-top:none;">' % LINEA)
    A('<table cellpadding="0" cellspacing="0" border="0" width="100%">')

    # ── Tarjetas resumen ──
    A('<tr><td style="padding-bottom:4px;"><table cellpadding="0" cellspacing="0" border="0" width="100%"><tr>')
    A(card(str(d['n_sem']), 'Cotizadas', AZUL))
    A(card(str(d['n_adj']), 'Adjudicadas', VERDE))
    A(card(str(d['n_fac']), 'Facturadas', '#15602B'))
    A(card(str(d['n_per']), 'Perdidas', ROJO))
    A(card(str(d['n_pend']), 'Sin cerrar (mes)', ORO))
    A('</tr></table></td></tr>')

    A('<tr><td style="padding:2px 0 6px 0;font-size:13px;color:#3D4C63;line-height:1.7;">'
      'Se cotizaron <b>%s</b> esta semana. De lo que se cerró, entraron <b>%s</b>. '
      'Quedan <b>%s</b> en cotizaciones vencidas que nadie ha marcado como ganadas o perdidas.'
      '</td></tr>' % (money(d['v_sem']), money(d['v_ganado_sem']), money(d['v_pend'])))

    # ── 1. Cotizaciones de la semana ──
    A(titulo('Cotizaciones emitidas esta semana',
             '%d cotizaciones por %s, agrupadas por vendedor' % (d['n_sem'], money(d['v_sem']))))
    if not d['semana']:
        A('<tr><td style="font-size:13px;color:%s;padding:10px 0;">No se emitieron cotizaciones '
          'entre el lunes y el viernes.</td></tr>' % GRIS)
    else:
        for vend, filas in d['semana_por_vend']:
            sub = sum(base_sin_iva(c) for c in filas)
            A('<tr><td style="padding:14px 0 6px 0;font-size:13.5px;color:#0F1B2D;font-weight:bold;">'
              '%s <span style="color:%s;font-weight:normal;font-size:12.5px;">&middot; %d cotizaciones '
              '&middot; %s</span></td></tr>' % (esc(vend), GRIS, len(filas), money(sub)))
            A('<tr><td><table cellpadding="0" cellspacing="0" border="0" width="100%%" '
              'style="border:1px solid %s;border-radius:3px;background:#ffffff;">' % LINEA)
            A('<tr>' + th('Cotización') + th('Fecha') + th('Cliente') +
              th('Valor', 'right') + th('Estado') + '</tr>')
            for c in filas:
                fd = parse_fecha(c['fecha'])
                extra_estado = chip(c['estado'])
                if c['estado'] == 'Rechazada' and c['motivoRechazo']:
                    extra_estado += ('<div style="font-size:10.5px;color:%s;padding-top:3px;">%s</div>'
                                     % (ROJO, esc(c['motivoRechazo'])))
                if c['estado'] == 'Facturada' and c['factura']:
                    extra_estado += ('<div style="font-size:10.5px;color:%s;padding-top:3px;">Fact. %s</div>'
                                     % (GRIS, esc(c['factura'])))
                if c['estado'] == 'Adjudicada' and c['valorAdjudicado'] is not None:
                    extra_estado += ('<div style="font-size:10.5px;color:%s;padding-top:3px;">Adj. %s</div>'
                                     % (VERDE, money(c['valorAdjudicado'])))
                A('<tr>' +
                  td('<b style="color:%s;">%s</b>' % (AZUL, esc(c['id']))) +
                  td('<span style="color:%s;">%s</span>' % (GRIS, fecha_corta(fd) if fd else '—')) +
                  td(esc(c['cliente'] or '—')) +
                  td(money(base_sin_iva(c)), 'right', 'white-space:nowrap;') +
                  td(extra_estado) + '</tr>')
            A('</table></td></tr>')

    # ── 2. Pendientes de cerrar ──
    A(titulo('⏰ Pendientes de cerrar — acumulado de %s' % d['mes_nombre'],
             'Cotizaciones de este mes que ya vencieron y siguen sin estado final. '
             'Esto es lo que hay que actualizar.'))
    if not d['pendientes']:
        A('<tr><td style="font-size:13px;color:%s;padding:10px 0;">Nada pendiente. '
          'Toda cotización vencida tiene su estado al día. 👏</td></tr>' % VERDE)
    else:
        A('<tr><td style="padding:6px 0 10px 0;">'
          '<table cellpadding="0" cellspacing="0" border="0" width="100%%" '
          'style="background:#FDF6E7;border-left:3px solid %s;border-radius:0 3px 3px 0;">'
          '<tr><td style="padding:12px 16px;font-size:13px;color:#3D4C63;line-height:1.6;">'
          '<b style="color:#9A6A05;">%d cotizaciones por %s</b> vencieron y siguen figurando como abiertas. '
          'No se sabe si se ganaron o se perdieron, y no cuentan en ningún indicador.<br>'
          'Se cierran en un clic desde la plataforma: <b>Cotizaciones &rarr; Base de Datos &rarr; filtro '
          '⏰ Sin cerrar</b>.'
          '</td></tr></table></td></tr>' % (ORO, d['n_pend'], money(d['v_pend'])))
        for vend, filas in d['pend_por_vend']:
            sub = sum(base_sin_iva(c) for c in filas)
            A('<tr><td style="padding:14px 0 6px 0;font-size:13.5px;color:#0F1B2D;font-weight:bold;">'
              '%s <span style="color:%s;font-weight:normal;font-size:12.5px;">&middot; %d por cerrar '
              '&middot; %s</span></td></tr>' % (esc(vend), GRIS, len(filas), money(sub)))
            A('<tr><td><table cellpadding="0" cellspacing="0" border="0" width="100%%" '
              'style="border:1px solid %s;border-radius:3px;background:#ffffff;">' % LINEA)
            A('<tr>' + th('Cotización') + th('Fecha') + th('Cliente') +
              th('Valor', 'right') + th('Vencida hace', 'right') + '</tr>')
            for c in filas[:12]:
                fd = parse_fecha(c['fecha'])
                dv = c['_dv']
                col = ROJO if dv >= 30 else ('#9A6A05' if dv >= 8 else GRIS)
                A('<tr>' +
                  td('<b style="color:%s;">%s</b>' % (AZUL, esc(c['id']))) +
                  td('<span style="color:%s;">%s</span>' % (GRIS, fecha_corta(fd) if fd else '—')) +
                  td(esc(c['cliente'] or '—')) +
                  td(money(base_sin_iva(c)), 'right', 'white-space:nowrap;') +
                  td('<b style="color:%s;">%d días</b>' % (col, dv), 'right') + '</tr>')
            if len(filas) > 12:
                A('<tr><td colspan="5" style="font-size:12px;color:%s;padding:8px 10px;">'
                  'y %d más &mdash; la lista completa está en el Excel adjunto</td></tr>'
                  % (GRIS, len(filas) - 12))
            A('</table></td></tr>')

    # ── 3. Adjudicadas sin facturar ──
    A(titulo('💳 Adjudicadas que faltan por facturar',
             'Negocios de %s ya ganados que todavía no tienen factura registrada' % d['mes_nombre']))
    if not d['sin_facturar']:
        A('<tr><td style="font-size:13px;color:%s;padding:10px 0;">Todo lo adjudicado está facturado.'
          '</td></tr>' % VERDE)
    else:
        A('<tr><td><table cellpadding="0" cellspacing="0" border="0" width="100%%" '
          'style="border:1px solid %s;border-radius:3px;background:#ffffff;">' % LINEA)
        A('<tr>' + th('Cotización') + th('Cliente') + th('Vendedor') +
          th('Adjudicado', 'right') + th('Adjudicada el', 'right') + '</tr>')
        for c in d['sin_facturar'][:15]:
            ad = parse_fecha(c['adjudicadaAt'])
            dias = (d['hoy'] - ad).days if ad else None
            txt = ('%s <span style="color:%s;">(%d días)</span>' % (fecha_corta(ad), GRIS, dias)) if ad else '—'
            A('<tr>' +
              td('<b style="color:%s;">%s</b>' % (AZUL, esc(c['id']))) +
              td(esc(c['cliente'] or '—')) +
              td('<span style="color:%s;font-size:12px;">%s</span>' % (GRIS, esc(vendedor_de(c)))) +
              td('<b>%s</b>' % money(valor_ganado(c)), 'right', 'white-space:nowrap;') +
              td(txt, 'right') + '</tr>')
        A('</table></td></tr>')
        A('<tr><td style="padding:8px 0;font-size:12.5px;color:%s;">%d por facturar, '
          '<b style="color:#0F1B2D;">%s</b> en total.</td></tr>'
          % (GRIS, len(d['sin_facturar']), money(sum(valor_ganado(c) for c in d['sin_facturar']))))

    # ── 4. Resumen por vendedor ──
    A(titulo('Resumen por vendedor', 'Semana del %s al %s' % (fecha_corta(lunes), fecha_corta(viernes))))
    A('<tr><td><table cellpadding="0" cellspacing="0" border="0" width="100%%" '
      'style="border:1px solid %s;border-radius:3px;background:#ffffff;">' % LINEA)
    A('<tr>' + th('Vendedor') + th('Cotizadas', 'right') + th('Valor cotizado', 'right') +
      th('Ganadas', 'right') + th('Valor ganado', 'right') + th('Sin cerrar', 'right') + '</tr>')
    for r in d['resumen_vend']:
        A('<tr>' +
          td('<b>%s</b>' % esc(r['vend'])) +
          td(str(r['n']), 'right') +
          td(money(r['v']), 'right', 'white-space:nowrap;') +
          td('<b style="color:%s;">%s</b>' % (VERDE, r['gn']) if r['gn'] else '<span style="color:%s;">0</span>' % GRIS, 'right') +
          td(money(r['gv']) if r['gn'] else '<span style="color:%s;">&mdash;</span>' % GRIS, 'right', 'white-space:nowrap;') +
          td(('<b style="color:%s;">%d &middot; %s</b>' % (ORO, r['pn'], money(r['pv']))) if r['pn']
             else '<span style="color:%s;">0</span>' % GRIS, 'right', 'white-space:nowrap;') + '</tr>')
    A('<tr style="background:#F6F8FB;">' +
      td('<b>Total</b>', 'left', 'border-bottom:none;border-top:2px solid %s;' % LINEA) +
      td('<b>%d</b>' % d['n_sem'], 'right', 'border-bottom:none;border-top:2px solid %s;' % LINEA) +
      td('<b>%s</b>' % money(d['v_sem']), 'right', 'border-bottom:none;border-top:2px solid %s;white-space:nowrap;' % LINEA) +
      td('<b>%d</b>' % d['n_ganadas_sem'], 'right', 'border-bottom:none;border-top:2px solid %s;' % LINEA) +
      td('<b>%s</b>' % money(d['v_ganado_sem']), 'right', 'border-bottom:none;border-top:2px solid %s;white-space:nowrap;' % LINEA) +
      td('<b>%d &middot; %s</b>' % (d['n_pend'], money(d['v_pend'])), 'right',
         'border-bottom:none;border-top:2px solid %s;white-space:nowrap;' % LINEA) + '</tr>')
    A('</table></td></tr>')

    # ── Pie ──
    A('<tr><td style="padding:24px 0 0 0;border-top:1px solid %s;margin-top:20px;">'
      '<div style="font-size:11.5px;color:%s;line-height:1.7;padding-top:14px;">'
      '<b>Cómo leerlo.</b> "Cotizadas" son las emitidas entre lunes y viernes de esta semana. '
      '"Adjudicadas" y "Facturadas" son el estado en que están hoy. "Sin cerrar" son las cotizaciones '
      'de %s que ya vencieron y siguen abiertas &mdash; acumuladas del mes, no de meses anteriores.<br>'
      '<b>Valores sin IVA</b>, para que sean comparables con el valor adjudicado.<br>'
      '<b>Informe automático</b> generado desde la plataforma ENERGY el %s a las %s. '
      'Es solo informativo: no requiere respuesta.'
      '</div></td></tr>' % (LINEA, GRIS, d['mes_nombre'], fecha_larga(d['hoy']), d['hora']))

    A('</table></td></tr>')
    A('</table></td></tr></table></div>')
    return '\n'.join(H)


# ─── EXCEL ────────────────────────────────────────────────────────────────────

def build_excel(d):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print('⚠️  openpyxl no disponible: se envía sin adjunto')
        return None

    wb = Workbook()
    cab_fill = PatternFill('solid', fgColor='1A3A8F')
    cab_font = Font(color='FFFFFF', bold=True, size=10)

    def hoja(titulo_h, cols, filas, primera=False):
        ws = wb.active if primera else wb.create_sheet()
        ws.title = titulo_h
        ws.append(cols)
        for i, _ in enumerate(cols, 1):
            cel = ws.cell(row=1, column=i)
            cel.fill = cab_fill
            cel.font = cab_font
            cel.alignment = Alignment(horizontal='center', vertical='center')
        for f in filas:
            ws.append(f)
        anchos = [14, 12, 30, 20, 16, 16, 14, 26, 12]
        for i, a in enumerate(anchos[:len(cols)], 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = a
        ws.freeze_panes = 'A2'
        if filas:
            ws.auto_filter.ref = 'A1:%s%d' % (ws.cell(row=1, column=len(cols)).column_letter, len(filas) + 1)
        return ws

    hoja('Semana', ['Cotización', 'Fecha', 'Cliente', 'Vendedor', 'Valor sin IVA', 'Adjudicado',
                    'Estado', 'Motivo rechazo', 'Factura'],
         [[c['id'], c['fecha'], c['cliente'], vendedor_de(c), round(base_sin_iva(c)),
           (round(float(c['valorAdjudicado'])) if c['valorAdjudicado'] is not None else ''),
           c['estado'], c['motivoRechazo'], c['factura']] for c in d['semana']],
         primera=True)

    hoja('Pendientes de cerrar', ['Cotización', 'Fecha', 'Cliente', 'Vendedor', 'Valor sin IVA',
                                  'Vence', 'Días vencida', 'Estado'],
         [[c['id'], c['fecha'], c['cliente'], vendedor_de(c), round(base_sin_iva(c)),
           c['fechaVenc'], c['_dv'], c['estado']] for c in d['pendientes']])

    hoja('Sin facturar', ['Cotización', 'Fecha', 'Cliente', 'Vendedor', 'Valor adjudicado',
                          'Adjudicada el', 'Estado'],
         [[c['id'], c['fecha'], c['cliente'], vendedor_de(c), round(valor_ganado(c)),
           c['adjudicadaAt'], c['estado']] for c in d['sin_facturar']])

    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── CORREO ───────────────────────────────────────────────────────────────────

def get_access_token(tenant_id, client_id, client_secret):
    url = 'https://login.microsoftonline.com/' + tenant_id + '/oauth2/v2.0/token'
    data = urllib.parse.urlencode({
        'grant_type': 'client_credentials', 'client_id': client_id,
        'client_secret': client_secret, 'scope': 'https://graph.microsoft.com/.default'
    }).encode()
    req = urllib.request.Request(url, data=data, method='POST')
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())['access_token']


def send_email(token, sender, to_list, cc_list, subject, html_body,
               attachment_bytes=None, attachment_name=None):
    """Igual que el de los otros informes, pero con COPIA (cc)."""
    msg = {
        'subject': subject,
        'body': {'contentType': 'HTML', 'content': html_body},
        'toRecipients': [{'emailAddress': {'address': r.strip()}} for r in to_list if r.strip()],
    }
    cc = [{'emailAddress': {'address': r.strip()}} for r in (cc_list or []) if r.strip()]
    if cc:
        msg['ccRecipients'] = cc
    if attachment_bytes and attachment_name:
        msg['attachments'] = [{
            '@odata.type': '#microsoft.graph.fileAttachment',
            'name': attachment_name,
            'contentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'contentBytes': base64.b64encode(attachment_bytes).decode('utf-8')
        }]
    payload = json.dumps({'message': msg, 'saveToSentItems': True}).encode('utf-8')
    url = 'https://graph.microsoft.com/v1.0/users/' + sender + '/sendMail'
    req = urllib.request.Request(url, data=payload, method='POST',
                                 headers={'Authorization': 'Bearer ' + token,
                                          'Content-Type': 'application/json'})
    with urllib.request.urlopen(req) as resp:
        print('✅ Correo enviado (HTTP %s)' % resp.status)


# ─── ARMADO ───────────────────────────────────────────────────────────────────

GANADO = ('Adjudicada', 'Facturada', 'Aprobada')
PERDIDO = ('Rechazada', 'Vencida', 'Anulada')


def construir(cots, hoy):
    # Semana: lunes a viernes de la semana en que cae `hoy`
    lunes = hoy - datetime.timedelta(days=hoy.weekday())
    viernes = lunes + datetime.timedelta(days=4)

    semana = [c for c in cots
              if (parse_fecha(c['fecha']) or datetime.date(1900, 1, 1)) >= lunes
              and (parse_fecha(c['fecha']) or datetime.date(2100, 1, 1)) <= viernes]
    semana.sort(key=lambda c: (vendedor_de(c), c['fecha'], c['id']))

    # Pendientes de cerrar: abiertas y vencidas, ACUMULADAS DEL MES EN CURSO.
    # No arrastra el histórico de meses anteriores — el informe es del mes, y así
    # el número es perseguible en vez de una montaña que nadie ataca.
    # Solo de la plataforma: las del LIBRO histórico no se editan aquí.
    mes_actual = hoy.strftime('%Y-%m')
    pendientes = []
    for c in cots:
        if c['fuente'] and c['fuente'] != 'Plataforma':
            continue
        if c['estado'] not in ESTADOS_ABIERTOS:
            continue
        if (c['fecha'] or '')[:7] != mes_actual:
            continue
        dv = dias_vencida(c, hoy)
        if dv:
            c['_dv'] = dv
            pendientes.append(c)
    pendientes.sort(key=lambda c: -c['_dv'])

    # Adjudicadas sin factura registrada — también acotadas al mes en curso, para
    # no arrastrar cotizaciones viejas sin fecha de adjudicación, que son ruido.
    sin_facturar = [c for c in cots
                    if c['estado'] == 'Adjudicada' and not c['factura']
                    and ((c['adjudicadaAt'] or '')[:7] == mes_actual
                         or (c['fecha'] or '')[:7] == mes_actual)]
    sin_facturar.sort(key=lambda c: -valor_ganado(c))

    def agrupar(lista):
        g = {}
        for c in lista:
            g.setdefault(vendedor_de(c), []).append(c)
        return sorted(g.items(), key=lambda kv: -sum(base_sin_iva(x) for x in kv[1]))

    ganadas_sem = [c for c in semana if c['estado'] in GANADO]
    perdidas_sem = [c for c in semana if c['estado'] in PERDIDO]

    # Resumen por vendedor: mezcla la semana con los pendientes de ese vendedor
    vends = set(vendedor_de(c) for c in semana) | set(vendedor_de(c) for c in pendientes)
    resumen = []
    for v in vends:
        s = [c for c in semana if vendedor_de(c) == v]
        p = [c for c in pendientes if vendedor_de(c) == v]
        g = [c for c in s if c['estado'] in GANADO]
        resumen.append({
            'vend': v, 'n': len(s), 'v': sum(base_sin_iva(c) for c in s),
            'gn': len(g), 'gv': sum(valor_ganado(c) for c in g),
            'pn': len(p), 'pv': sum(base_sin_iva(c) for c in p),
        })
    resumen.sort(key=lambda r: -r['v'])

    return {
        'hoy': hoy, 'lunes': lunes, 'viernes': viernes,
        'mes_nombre': ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto',
                       'septiembre', 'octubre', 'noviembre', 'diciembre'][hoy.month - 1],
        'hora': (now_co()).strftime('%I:%M %p').lstrip('0'),
        'semana': semana, 'semana_por_vend': agrupar(semana),
        'pendientes': pendientes, 'pend_por_vend': agrupar(pendientes),
        'sin_facturar': sin_facturar,
        'resumen_vend': resumen,
        'n_sem': len(semana), 'v_sem': sum(base_sin_iva(c) for c in semana),
        'n_adj': len([c for c in semana if c['estado'] == 'Adjudicada']),
        'n_fac': len([c for c in semana if c['estado'] == 'Facturada']),
        'n_per': len(perdidas_sem),
        'n_ganadas_sem': len(ganadas_sem),
        'v_ganado_sem': sum(valor_ganado(c) for c in ganadas_sem),
        'n_pend': len(pendientes), 'v_pend': sum(base_sin_iva(c) for c in pendientes),
    }


if __name__ == '__main__':
    hoy = now_co().date()
    print('📥 Leyendo cotizaciones de Supabase...')
    cots = cargar_cotizaciones()
    d = construir(cots, hoy)

    print('   Semana %s a %s: %d cotizaciones · %d pendientes de cerrar · %d sin facturar'
          % (d['lunes'], d['viernes'], d['n_sem'], d['n_pend'], len(d['sin_facturar'])))

    html = build_html(d)
    xlsx = build_excel(d)
    asunto = ('📊 Informe Comercial Semanal · %s al %s'
              % (fecha_corta(d['lunes']), fecha_corta(d['viernes'])))
    nombre_xlsx = 'Comercial_Semana_%s.xlsx' % d['viernes'].isoformat()

    # Modo prueba: escribe los archivos en vez de enviar el correo
    test_out = os.environ.get('TEST_OUT', '').strip()
    if test_out:
        p = os.path.join(test_out, 'informe_semanal_test.html')
        with open(p, 'w', encoding='utf-8') as f:
            f.write(html)
        print('🧪 HTML de prueba: ' + p)
        if xlsx:
            px = os.path.join(test_out, nombre_xlsx)
            with open(px, 'wb') as f:
                f.write(xlsx)
            print('🧪 Excel de prueba: ' + px)
        raise SystemExit(0)

    tenant = os.environ.get('MS_TENANT_ID', '').strip()
    cid = os.environ.get('MS_CLIENT_ID', '').strip()
    secret = os.environ.get('MS_CLIENT_SECRET', '').strip()
    sender = os.environ.get('SENDER_EMAIL', '').strip()
    to_list = [r.strip() for r in os.environ.get('RECIPIENT_EMAILS', '').split(',') if r.strip()]
    cc_list = [r.strip() for r in os.environ.get('CC_EMAILS', '').split(',') if r.strip()]

    if not (tenant and cid and secret and sender and to_list):
        raise SystemExit('❌ Faltan credenciales o destinatarios (MS_*, SENDER_EMAIL, RECIPIENT_EMAILS)')

    print('📧 Enviando a %s (copia: %s)' % (', '.join(to_list), ', '.join(cc_list) or 'nadie'))
    token = get_access_token(tenant, cid, secret)
    send_email(token, sender, to_list, cc_list, asunto, html, xlsx, nombre_xlsx if xlsx else None)
