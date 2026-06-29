# -*- coding: utf-8 -*-
"""
Extrae del LIBRO COTIZACIONES.xlsx las cotizaciones de los ultimos 3 meses
(ventana movil 90 dias desde HOY) y refresca data/cotizaciones_historicas.json.

Reglas acordadas:
  - Periodo: fecha de generacion >= HOY-90d  y <= HOY
  - Destino: cotizaciones_historicas.json (mismo esquema de 17 campos)
  - Conflictos: gana la plataforma web -> ids presentes en cotizaciones.json NO se agregan
  - No destructivo: se conserva todo lo anterior a la ventana; solo se reemplazan
    los ids que SI logramos extraer del LIBRO en la ventana.

Modo:  python extraer_libro_3meses.py            -> DRY RUN (no escribe nada)
       python extraer_libro_3meses.py --write     -> hace backup y escribe
"""
import openpyxl, json, os, sys, datetime, re

LIBRO   = r"C:\Users\Lenovo\OneDrive\Escritorio\Energy_bot\LIBRO COTIZACIONES .xlsx"
DATA    = r"C:\Users\Lenovo\OneDrive\Escritorio\Energy_bot\plataforma-eyg\data"
HIST    = os.path.join(DATA, "cotizaciones_historicas.json")
WEB     = os.path.join(DATA, "cotizaciones.json")
PREVIEW = r"C:\Users\Lenovo\AppData\Local\Temp\eyg_cotiz\historicas_preview.json"
os.makedirs(os.path.dirname(PREVIEW), exist_ok=True)

HOY    = datetime.date.today()                # se ajusta solo a la fecha de la corrida
CUTOFF = HOY - datetime.timedelta(days=90)    # ventana movil de 90 dias
SALTAR = {"CONSECUTIVOS", "FORMATO"}

def norm_id(s):
    return re.sub(r"[\s\-_]", "", str(s or "")).upper()

def num(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(".", "").replace(",", ".")) if re.search(r"[.,]\d{1,2}$", str(v)) else float(str(v).replace(",", ""))
    except Exception:
        return 0

def parse_sheet(name, ws):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return None
    top = rows[:9]
    fecha = cliente = None
    fecha_venc = None
    contacto = correo = None
    # fecha de generacion + cliente por etiqueta
    fechas_dt = []
    for row in top:
        for ci, val in enumerate(row):
            if isinstance(val, datetime.datetime):
                fechas_dt.append(val)
            s = str(val).upper().strip() if val is not None else ""
            if "GENERAC" in s:
                for c2 in range(ci + 1, len(row)):
                    if isinstance(row[c2], datetime.datetime):
                        fecha = row[c2]; break
            if "VENCIMIEN" in s:
                for c2 in range(ci + 1, len(row)):
                    if isinstance(row[c2], datetime.datetime):
                        fecha_venc = row[c2]; break
            if s == "CLIENTE":
                for c2 in range(ci + 1, min(ci + 4, 6)):   # acotado: no invadir col FECHA (idx 6)
                    if row[c2] not in (None, ""):
                        cliente = row[c2]; break
            if s == "CONTACTO":
                for c2 in range(ci + 1, min(ci + 4, 6)):
                    if row[c2] not in (None, ""):
                        contacto = row[c2]; break
            if s == "CORREO":
                for c2 in range(ci + 1, min(ci + 4, 6)):
                    if row[c2] not in (None, ""):
                        correo = row[c2]; break
    if fecha is None and fechas_dt:
        fecha = min(fechas_dt)          # generacion suele ser la menor (vencimiento es mayor)
    if fecha is None:
        return None
    # fila de encabezados de items
    hdr = None; cols = {}
    for ri, row in enumerate(rows[:12]):
        joined = " ".join(str(x).upper() for x in row if x)
        if "DESCRIPCION" in joined and "CANTIDAD" in joined:
            hdr = ri
            for ci, val in enumerate(row):
                s = str(val).upper().strip() if val is not None else ""
                if s.startswith("ITEM"): cols["item"] = ci
                elif "UNIDAD" in s or s == "UDM": cols["udm"] = ci
                elif "CANTIDAD" in s: cols["qty"] = ci
                elif "DESCRIPCION" in s: cols["desc"] = ci
                elif "TIEMPO" in s and "ENTREGA" in s: cols["tiempo"] = ci
                elif "DESVIACION" in s: cols["desv"] = ci
                elif "MARCA" in s: cols["marca"] = ci
                elif "VALOR UNITARIO" in s: cols["vunit"] = ci
                elif "VALOR TOTAL" in s: cols["vtotal"] = ci
            break
    if hdr is None:
        return None
    # total de la cotizacion (label TOTAL, no SUBTOTAL) bajo los items
    quote_total = 0; quote_sub = 0
    observ = ""; forma_pago = ""; sitio = ""; validez = ""
    for ri2, row in enumerate(rows[hdr + 1:], start=hdr + 1):
        for ci, val in enumerate(row):
            s = str(val).upper().strip() if val is not None else ""
            if s in ("TOTAL", "TOTAL:", "GRAN TOTAL", "VALOR TOTAL OFERTA"):
                for c2 in range(ci + 1, len(row)):
                    if isinstance(row[c2], (int, float)) and row[c2]:
                        quote_total = float(row[c2]); break
            if s in ("SUBTOTAL", "SUB TOTAL", "SUBTOTAL:"):
                for c2 in range(ci + 1, len(row)):
                    if isinstance(row[c2], (int, float)) and row[c2]:
                        quote_sub = float(row[c2]); break
            # OBSERVACIONES: texto en la celda a la derecha (misma fila)
            if s == "OBSERVACIONES":
                for c2 in range(ci + 1, len(row)):
                    if row[c2] not in (None, ""):
                        observ = str(row[c2]).strip(); break
            # CONDICIONES COMERCIALES: etiquetas en esta fila, valores en la siguiente
            if "FORMA DE PAGO" in s:
                nxt = rows[ri2 + 1] if ri2 + 1 < len(rows) else []
                col_sitio = col_val = None
                for cj, v2 in enumerate(row):
                    sj = str(v2).upper().strip() if v2 is not None else ""
                    if "FORMA DE PAGO" in sj and cj < len(nxt):
                        forma_pago = str(nxt[cj]).strip() if nxt[cj] not in (None, "") else ""
                    if "SITIO DE ENTREGA" in sj and cj < len(nxt):
                        sitio = str(nxt[cj]).strip() if nxt[cj] not in (None, "") else ""
                    if "VALIDEZ" in sj and cj < len(nxt):
                        validez = str(nxt[cj]).strip() if nxt[cj] not in (None, "") else ""
    # items
    ci_item = cols.get("item", 0)
    items = []
    started = False
    for row in rows[hdr + 1:]:
        raw_item = row[ci_item] if ci_item < len(row) else None
        is_item = raw_item is not None and str(raw_item).strip().replace(".0", "").isdigit()
        desc = row[cols["desc"]] if "desc" in cols and cols["desc"] < len(row) else None
        if is_item:
            started = True
            def cell(key):
                ci = cols.get(key)
                if ci is None or ci >= len(row) or row[ci] in (None, ""):
                    return ""
                return str(row[ci]).strip()
            items.append({
                "item": int(float(raw_item)),
                "desc": str(desc).strip() if desc else "",
                "udm": str(row[cols["udm"]]).strip() if cols.get("udm") is not None and row[cols["udm"]] else "",
                "qty": num(row[cols["qty"]]) if cols.get("qty") is not None else 0,
                "v_unit": num(row[cols["vunit"]]) if cols.get("vunit") is not None else 0,
                "v_total": num(row[cols["vtotal"]]) if cols.get("vtotal") is not None else 0,
                "marca": cell("marca"),
                "tiempo": cell("tiempo"),
                "desv": cell("desv"),
            })
        elif started:
            # primera fila sin numero de item tras empezar -> fin de la tabla
            if all((x in (None, "")) for x in (row[:4] if len(row) >= 4 else row)):
                break
            # si hay texto pero no item, puede ser nota/total -> cortamos
            break
    return {
        "id": name.strip(),
        "fecha": fecha.date(),
        "fecha_venc": fecha_venc.date() if fecha_venc else None,
        "cliente": str(cliente).strip() if cliente else "",
        "contacto": str(contacto).strip() if contacto else "",
        "correo": str(correo).strip() if correo else "",
        "total": quote_total, "subtotal": quote_sub,
        "observaciones": observ, "forma_pago": forma_pago,
        "sitio_entrega": sitio, "validez": validez,
        "items": items,
    }

def main():
    write = "--write" in sys.argv
    wb = openpyxl.load_workbook(LIBRO, read_only=True, data_only=True)
    web = json.load(open(WEB, encoding="utf-8"))
    web_ids = {norm_id(c.get("id")) for c in web}
    hist = json.load(open(HIST, encoding="utf-8"))

    extraidas = []          # cotizaciones (objeto) en ventana
    saltadas_web = 0
    fuera_ventana = 0
    sin_parse = 0
    total_hojas = 0

    for name in wb.sheetnames:
        if name.strip().upper() in SALTAR:
            continue
        total_hojas += 1
        try:
            ws = wb[name]
            q = parse_sheet(name, ws)
        except Exception as e:
            sin_parse += 1
            continue
        if q is None:
            sin_parse += 1
            continue
        if not (CUTOFF <= q["fecha"] <= HOY):
            fuera_ventana += 1
            continue
        if norm_id(q["id"]) in web_ids:
            saltadas_web += 1
            continue
        extraidas.append(q)

    # construir lineas nuevas: 17 campos originales (mismo orden) + 3 aditivos
    def make_line(q, fstr, it):
        return {"id": q["id"], "fecha": fstr, "cliente": q["cliente"],
                "item": it["item"], "desc": it["desc"], "udm": it["udm"], "qty": it["qty"],
                "v_unit": it["v_unit"], "v_total": it["v_total"],
                "subtotal": q["subtotal"], "total": q["total"], "forma_pago": q["forma_pago"],
                "proveedor": "", "costo": None, "estado": "", "factura": "", "clasificacion": "",
                # --- campos aditivos nuevos ---
                "observaciones": q["observaciones"], "sitio_entrega": q["sitio_entrega"],
                "validez": q["validez"],
                # --- recuperados del LIBRO (cabecera + item) ---
                "fecha_venc": q.get("fecha_venc").isoformat() if q.get("fecha_venc") else "",
                "contacto": q.get("contacto", ""), "correo": q.get("correo", ""),
                "marca": it.get("marca", ""), "tiempo": it.get("tiempo", ""),
                "desv": it.get("desv", "")}

    nuevas = []
    for q in extraidas:
        fstr = q["fecha"].isoformat()
        items = q["items"] or [{"item": 0, "desc": "", "udm": "", "qty": 0, "v_unit": 0, "v_total": 0}]
        for it in items:
            nuevas.append(make_line(q, fstr, it))

    extraidos_norm = {norm_id(q["id"]) for q in extraidas}

    def fecha_valida(r):
        try:
            return datetime.date.fromisoformat(str(r.get("fecha"))[:10])
        except Exception:
            return None

    conservadas = []
    reemplazadas_ids = set()
    for r in hist:
        f = fecha_valida(r)
        nid = norm_id(r.get("id"))
        if f is not None and f >= CUTOFF and nid in extraidos_norm:
            reemplazadas_ids.add(nid)
            continue   # se reemplaza por la version fresca del LIBRO
        conservadas.append(r)

    final = conservadas + nuevas

    # ---- resumen ----
    con_precio = sum(1 for q in extraidas if q["total"] or any(i["v_unit"] for i in q["items"]))
    print("=" * 60)
    print(f"VENTANA: {CUTOFF} a {HOY}  (90 dias)")
    print(f"Hojas de cotizacion en el LIBRO        : {total_hojas}")
    print(f"  - fuera de la ventana (otra fecha)   : {fuera_ventana}")
    print(f"  - no parseables (layout viejo/vacio)  : {sin_parse}")
    print(f"  - saltadas porque ya estan en la web  : {saltadas_web}")
    print(f"  => COTIZACIONES extraidas en ventana  : {len(extraidas)}")
    print(f"       de ellas con algun precio/total  : {con_precio}")
    print(f"       lineas de item generadas         : {len(nuevas)}")
    print("-" * 60)
    print(f"Historico original (lineas)            : {len(hist)}")
    print(f"  - ids en ventana reemplazados        : {len(reemplazadas_ids)}")
    print(f"  - conservadas (resto del historico)  : {len(conservadas)}")
    print(f"  => HISTORICO FINAL (lineas)          : {len(final)}")
    print("=" * 60)
    con_obs = sum(1 for q in extraidas if q["observaciones"])
    con_fp  = sum(1 for q in extraidas if q["forma_pago"])
    print(f"Con observaciones: {con_obs}/{len(extraidas)} | con forma_pago: {con_fp}/{len(extraidas)}")
    print("-" * 60)
    print("Muestra de cotizaciones extraidas:")
    for q in sorted(extraidas, key=lambda x: x["fecha"])[-5:]:
        print(f"  {q['id']:12} {q['fecha']}  {str(q['cliente'])[:20]:20} items={len(q['items'])} total={q['total']:,.0f}")
        print(f"     forma_pago={q['forma_pago'][:20]!r} sitio={q['sitio_entrega'][:25]!r} validez={q['validez'][:12]!r}")
        print(f"     obs={q['observaciones'][:70]!r}")

    json.dump(final, open(PREVIEW, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"\nPreview escrito en: {PREVIEW}")

    if write:
        import shutil
        bak = HIST + ".bak_" + HOY.isoformat()
        shutil.copy2(HIST, bak)
        json.dump(final, open(HIST, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        print(f"\n*** ESCRITO en {HIST}")
        print(f"*** Backup en  {bak}")
    else:
        print("\n(DRY RUN - no se modifico el archivo real. Use --write para aplicar.)")

if __name__ == "__main__":
    main()
