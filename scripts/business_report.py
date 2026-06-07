"""
ENERGY — Informe Comercial Diario + Excel de Datos
===================================================
Une las cotizaciones históricas (Excel viejo), las de la plataforma (GitHub)
y los datos de Supabase (remisiones, plan de compras, OCs, proveedores).
Calcula KPIs (tasa de conversión, pipeline, top clientes, margen) y envía
un correo HTML a Andrea con el archivo Datos_EYG.xlsx adjunto.

Ejecutado por GitHub Actions L-V 7:00 PM Colombia (00:00 UTC mar-sáb).
"""

import json
import io
import os
import sys
import base64
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timezone, timedelta
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

DAYS_ES   = {'Monday':'Lunes','Tuesday':'Martes','Wednesday':'Miércoles',
             'Thursday':'Jueves','Friday':'Viernes','Saturday':'Sábado','Sunday':'Domingo'}
MONTHS_ES = {'January':'enero','February':'febrero','March':'marzo','April':'abril',
             'May':'mayo','June':'junio','July':'julio','August':'agosto',
             'September':'septiembre','October':'octubre','November':'noviembre','December':'diciembre'}

SUPABASE_URL_DEFAULT = 'https://juprjevxkcitqpsnemto.supabase.co'
SUPABASE_KEY_DEFAULT = 'sb_publishable_zZrmpmvqbz4AJCGHRHQ8Xw_8tnf5ObM'

ESTADOS_GANADA  = {'APROBADA', 'FACTURADA', 'ADJUDICADA', 'ACEPTADA'}
ESTADOS_PERDIDA = {'RECHAZADA', 'VENCIDA', 'ANULADA', 'CERRADA'}

REPO_ROOT = os.path.join(os.path.dirname(__file__), '..')


def now_co():
    return datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=-5)))


def money(n):
    try:
        return '$ ' + format(int(round(float(n or 0))), ',').replace(',', '.')
    except Exception:
        return '$ 0'


# ─── 1. CARGA DE DATOS ────────────────────────────────────────────────────────

def load_json(rel_path):
    p = os.path.join(REPO_ROOT, rel_path)
    try:
        with open(p, 'r', encoding='utf-8') as f:
            d = json.load(f)
            return d if isinstance(d, list) else []
    except Exception as e:
        print(f'⚠️  No se pudo leer {rel_path}: {e}')
        return []


def fetch_supabase(table, params=''):
    url_base = os.environ.get('SUPABASE_URL', SUPABASE_URL_DEFAULT).strip().rstrip('/')
    key = os.environ.get('SUPABASE_KEY', SUPABASE_KEY_DEFAULT).strip()
    url = f'{url_base}/rest/v1/{table}?limit=10000{params}'
    req = urllib.request.Request(url, headers={'apikey': key, 'Authorization': 'Bearer ' + key})
    try:
        with urllib.request.urlopen(req) as resp:
            rows = json.loads(resp.read())
            print(f'   Supabase {table}: {len(rows)} filas')
            return rows
    except Exception as e:
        print(f'⚠️  Supabase {table} error: {e}')
        return []


def unify_cotizaciones(historicas, plataforma):
    """Une ambas fuentes a nivel COTIZACIÓN: {id, fecha, cliente, estado, total, fuente}."""
    cots = {}
    # Históricas (línea por línea → agrupar)
    for r in historicas:
        cid = str(r.get('id') or '').strip()
        if not cid:
            continue
        c = cots.setdefault(cid, {'id': cid, 'fecha': '', 'cliente': '', 'estado': '',
                                  'total': 0, 'n_items': 0, 'fuente': 'Excel histórico',
                                  'venta': 0, 'costo': 0})
        c['fecha'] = str(r.get('fecha') or c['fecha'])[:10]
        c['cliente'] = r.get('cliente') or c['cliente']
        c['estado'] = r.get('estado') or c['estado']
        if r.get('total'):
            c['total'] = float(r.get('total') or 0)
        c['n_items'] += 1
        try:
            qty = float(r.get('qty') or 0)
            c['venta'] += qty * float(r.get('v_unit') or 0)
            c['costo'] += qty * float(r.get('costo') or 0)
        except Exception:
            pass
    # Plataforma (objeto por cotización) — pisa al histórico si el id se repite
    for c0 in plataforma:
        cid = str(c0.get('id') or '').strip()
        if not cid:
            continue
        venta = costo = 0
        for it in (c0.get('items') or []):
            try:
                qty = float(it.get('qty') or 0)
                venta += qty * float(it.get('precio') or 0)
                costo += qty * float(it.get('precioProveedor') or 0)
            except Exception:
                pass
        cots[cid] = {'id': cid, 'fecha': str(c0.get('fecha') or '')[:10],
                     'cliente': c0.get('cliente') or '', 'estado': c0.get('estado') or '',
                     'total': float(c0.get('total') or 0), 'n_items': len(c0.get('items') or []),
                     'fuente': 'Plataforma', 'venta': venta, 'costo': costo,
                     'updatedAt': str(c0.get('updatedAt') or ''), 'createdAt': str(c0.get('createdAt') or '')}
    return list(cots.values())


def clasifica(estado):
    e = str(estado or '').strip().upper()
    if e in ESTADOS_GANADA:
        return 'GANADA'
    if e in ESTADOS_PERDIDA:
        return 'PERDIDA'
    return 'ABIERTA'


# ─── 2. KPIs ──────────────────────────────────────────────────────────────────

def kpis_mes(cots, ym):
    mes = [c for c in cots if (c['fecha'] or '')[:7] == ym]
    ganadas = [c for c in mes if clasifica(c['estado']) == 'GANADA']
    abiertas = [c for c in mes if clasifica(c['estado']) == 'ABIERTA']
    return {
        'n': len(mes), 'monto': sum(c['total'] for c in mes),
        'n_ganadas': len(ganadas), 'monto_ganado': sum(c['total'] for c in ganadas),
        'conv': (len(ganadas) / len(mes) * 100) if mes else 0,
        'pipeline': sum(c['total'] for c in abiertas),
    }


def serie_meses(cots, n=6):
    hoy = now_co()
    out = []
    y, m = hoy.year, hoy.month
    for _ in range(n):
        ym = f'{y:04d}-{m:02d}'
        out.append((ym, kpis_mes(cots, ym)))
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return list(reversed(out))


def top_clientes(cots, year, top=5):
    acc = defaultdict(lambda: {'ganado': 0, 'n_cot': 0, 'n_ganadas': 0})
    for c in cots:
        if (c['fecha'] or '')[:4] != str(year):
            continue
        cli = (c['cliente'] or '—').strip().upper()
        acc[cli]['n_cot'] += 1
        if clasifica(c['estado']) == 'GANADA':
            acc[cli]['ganado'] += c['total']
            acc[cli]['n_ganadas'] += 1
    lista = sorted(acc.items(), key=lambda kv: -kv[1]['ganado'])
    return lista[:top]


def hoy_actividad(cots_plataforma, remisiones, planes, ocs, solicitudes):
    hoy = now_co().date().isoformat()
    act = {}
    act['cotiz'] = [c for c in cots_plataforma
                    if str(c.get('updatedAt') or c.get('createdAt') or '')[:10] == hoy
                    or str(c.get('fecha') or '')[:10] == hoy]
    act['remis'] = sorted({str(r.get('remision')) for r in remisiones if str(r.get('fecha') or '')[:10] == hoy})
    act['planes'] = sorted({str(p.get('cc')) for p in planes if str(p.get('created_at') or '')[:10] == hoy})
    act['ocs'] = [o for o in ocs if str(o.get('fecha') or '')[:10] == hoy]
    act['sols'] = [s for s in solicitudes if str(s.get('fecha') or s.get('createdAt') or '')[:10] == hoy
                   or ('SOL-' + hoy.replace('-', '')) in str(s.get('id') or '')]
    return act


# ─── 3. EXCEL DE DATOS ────────────────────────────────────────────────────────

HDR_FILL = None
HDR_FONT = None

def _sheet_from_rows(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center')
    for i, r in enumerate(rows, 2):
        for col, v in enumerate(r, 1):
            ws.cell(row=i, column=col, value=v)
    ws.freeze_panes = 'A2'
    if rows:
        ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows)+1}'
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
    return ws


def generate_excel(cots, lineas_hist, cots_plataforma, remisiones, planes, ocs, proveedores, solicitudes, serie):
    if not OPENPYXL_OK:
        return None
    global HDR_FILL, HDR_FONT
    HDR_FILL = PatternFill('solid', fgColor='0F2B5B')
    HDR_FONT = Font(bold=True, color='FFFFFF', size=10)

    wb = openpyxl.Workbook()
    # Hoja KPIs
    ws = wb.active
    ws.title = 'KPIs'
    ws['A1'] = 'E&G ENERGY GROUP — Datos al ' + now_co().strftime('%Y-%m-%d %H:%M')
    ws['A1'].font = Font(bold=True, size=13, color='0F2B5B')
    ws.append([])
    ws.append(['Mes', 'Cotizaciones', 'Monto cotizado', 'Ganadas', 'Monto ganado', '% Conversión', 'Pipeline abierto'])
    for c in ws[3]:
        c.fill = HDR_FILL; c.font = HDR_FONT
    for ym, k in serie:
        ws.append([ym, k['n'], round(k['monto']), k['n_ganadas'], round(k['monto_ganado']),
                   round(k['conv'], 1), round(k['pipeline'])])
    for col in 'ABCDEFG':
        ws.column_dimensions[col].width = 17

    # Cotizaciones (1 fila por cotización, ambas fuentes)
    _sheet_from_rows(wb, 'Cotizaciones',
        ['ID', 'Fecha', 'Cliente', 'Estado', 'Clasificación', 'Total', 'Ítems', 'Venta', 'Costo', 'Margen %', 'Fuente'],
        [[c['id'], c['fecha'], c['cliente'], c['estado'], clasifica(c['estado']),
          round(c['total']), c['n_items'], round(c['venta']), round(c['costo']),
          (round((1 - c['costo']/c['venta'])*100, 1) if c['venta'] and c['costo'] else ''),
          c['fuente']]
         for c in sorted(cots, key=lambda x: x['fecha'] or '', reverse=True)])

    # Líneas de cotización (histórico + plataforma)
    lin_rows = [[r.get('id'), r.get('fecha'), r.get('cliente'), r.get('item'), r.get('desc'),
                 r.get('udm'), r.get('qty'), r.get('v_unit'), r.get('proveedor'), r.get('costo'),
                 r.get('estado'), r.get('clasificacion'), 'Excel histórico'] for r in lineas_hist]
    for c0 in cots_plataforma:
        for it in (c0.get('items') or []):
            lin_rows.append([c0.get('id'), str(c0.get('fecha') or '')[:10], c0.get('cliente'),
                             it.get('id'), str(it.get('desc') or '')[:200], it.get('udm'), it.get('qty'),
                             it.get('precio'), it.get('proveedor'), it.get('precioProveedor'),
                             c0.get('estado'), '', 'Plataforma'])
    _sheet_from_rows(wb, 'Lineas_Cotizacion',
        ['ID Cot', 'Fecha', 'Cliente', 'Ítem', 'Descripción', 'UDM', 'Cant', 'Precio venta',
         'Proveedor', 'Costo prov', 'Estado', 'Clasificación', 'Fuente'], lin_rows)

    _sheet_from_rows(wb, 'Remisiones',
        ['Remisión', 'Fecha', 'Cliente', 'OC cliente', 'Ítem', 'Descripción', 'Cant', 'Marca', 'Creada por'],
        [[r.get('remision'), r.get('fecha'), r.get('cliente'), r.get('oc'), r.get('item'),
          r.get('descripcion'), r.get('cantidad'), r.get('marca'), r.get('created_by')] for r in remisiones])

    _sheet_from_rows(wb, 'Plan_Compras',
        ['CC', 'Cotización', 'Cliente', 'Fecha', 'Ítem', 'Descripción', 'UDM', 'Cant', 'Proveedor',
         'Costo unit', 'Pago', 'Sin OC', 'Creado por'],
        [[p.get('cc'), p.get('cotizacion'), p.get('cliente'), p.get('fecha'), p.get('item'),
          p.get('descripcion'), p.get('udm'), p.get('cantidad'), p.get('proveedor'),
          p.get('costo_unit'), p.get('pago'), ('SI' if p.get('sin_oc') else ''), p.get('created_by')] for p in planes])

    _sheet_from_rows(wb, 'OCs_Compra',
        ['# OC', 'CC', 'Cotización', 'Proveedor', 'Fecha', 'Subtotal', 'IVA', 'Retención', 'Total', 'Creada por'],
        [[o.get('oc'), o.get('cc'), o.get('cotizacion'), o.get('proveedor'), o.get('fecha'),
          o.get('subtotal'), o.get('iva'), o.get('retencion'), o.get('total'), o.get('created_by')] for o in ocs])

    _sheet_from_rows(wb, 'Proveedores',
        ['Nombre', 'Autoretenedor', 'NIT', 'Dirección', 'Ciudad', 'Teléfono', 'Contacto', 'Email'],
        [[p.get('nombre'), p.get('autoretenedor'), p.get('nit'), p.get('direccion'), p.get('ciudad'),
          p.get('telefono'), p.get('contacto'), p.get('email')] for p in proveedores])

    _sheet_from_rows(wb, 'Solicitudes',
        ['ID', 'Fecha', 'Cliente', 'Estado', 'Cotización'],
        [[s.get('id'), str(s.get('fecha') or s.get('createdAt') or '')[:10],
          s.get('cliente') or s.get('empresa'), s.get('estado'), s.get('cotizacionId') or ''] for s in solicitudes])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── 4. EMAIL HTML ────────────────────────────────────────────────────────────

def _card(num, lbl, bg, border, color):
    return (f"<td bgcolor='{bg}' style='background:{bg};border-left:4px solid {border};"
            f"padding:14px 6px;text-align:center;'>"
            f"<div style='font-size:26px;font-weight:900;color:{color};line-height:1;'>{num}</div>"
            f"<div style='font-size:10px;color:#666;margin-top:4px;text-transform:uppercase;letter-spacing:.5px;'>{lbl}</div></td>")


def generate_html(date_str, act, k_mes, serie, tops, margen_mes, totales):
    # Actividad de hoy
    cot_rows = ''.join(
        f"<tr><td style='padding:6px 10px;border-bottom:1px solid #eef1f8;'><b style='color:#0F2B5B'>{c.get('id')}</b></td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eef1f8;'>{c.get('cliente') or '—'}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eef1f8;'>{c.get('estado') or '—'}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eef1f8;text-align:right;'>{money(c.get('total'))}</td></tr>"
        for c in act['cotiz'][:15])
    hoy_html = (
        "<table width='100%' cellpadding='6' cellspacing='6'><tr>"
        + _card(len(act['cotiz']), 'Cotizaciones hoy', '#e8f0fe', '#1A3A8F', '#0F2B5B')
        + _card(len(act['sols']), 'Solicitudes hoy', '#fff8e1', '#E8A020', '#b7770d')
        + _card(len(act['remis']), 'Remisiones hoy', '#e6f4ea', '#2EAA4A', '#1B5E20')
        + _card(len(act['planes']), 'Planes compra hoy', '#f3e8ff', '#7B1FA2', '#4A148C')
        + _card(len(act['ocs']), 'OCs compra hoy', '#fce8e6', '#e53e3e', '#c0392b')
        + "</tr></table>")
    if cot_rows:
        hoy_html += ("<table width='100%' cellpadding='0' cellspacing='0' style='border-collapse:collapse;font-size:12px;margin-top:6px;'>"
                     "<thead><tr><th style='background:#1A3A8F;color:#fff;padding:6px 10px;text-align:left;font-size:10px;'>COTIZACIÓN</th>"
                     "<th style='background:#1A3A8F;color:#fff;padding:6px 10px;text-align:left;font-size:10px;'>CLIENTE</th>"
                     "<th style='background:#1A3A8F;color:#fff;padding:6px 10px;text-align:left;font-size:10px;'>ESTADO</th>"
                     "<th style='background:#1A3A8F;color:#fff;padding:6px 10px;text-align:right;font-size:10px;'>TOTAL</th></tr></thead>"
                     f"<tbody>{cot_rows}</tbody></table>")

    # Serie 6 meses
    serie_rows = ''.join(
        f"<tr><td style='padding:6px 10px;border-bottom:1px solid #eef1f8;font-weight:700;color:#0F2B5B;'>{ym}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eef1f8;text-align:center;'>{k['n']}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eef1f8;text-align:right;'>{money(k['monto'])}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eef1f8;text-align:center;'>{k['n_ganadas']}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eef1f8;text-align:right;'>{money(k['monto_ganado'])}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eef1f8;text-align:center;'>"
        f"<b style='color:{'#1B5E20' if k['conv']>=25 else ('#b7770d' if k['conv']>=10 else '#c0392b')};'>{k['conv']:.0f}%</b></td></tr>"
        for ym, k in serie)

    tops_rows = ''.join(
        f"<tr><td style='padding:6px 10px;border-bottom:1px solid #eef1f8;'><b>{i+1}. {cli.title()}</b></td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eef1f8;text-align:right;'>{money(d['ganado'])}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eef1f8;text-align:center;'>{d['n_ganadas']}/{d['n_cot']}</td></tr>"
        for i, (cli, d) in enumerate(tops))

    margen_html = ''
    if margen_mes is not None:
        margen_html = (f"<div style='background:#f8f9ff;border-radius:8px;padding:10px 14px;font-size:12px;color:#555;margin:10px 0;'>"
                       f"&#128176; Margen bruto estimado del mes (donde hay costo registrado): "
                       f"<b style='color:#0F2B5B;'>{margen_mes:.1f}%</b></div>")

    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><style>
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f4ff;margin:0;padding:16px;}}
  .wrap{{max-width:760px;margin:0 auto;box-shadow:0 8px 32px rgba(0,0,0,.15);border-radius:14px;overflow:hidden;}}
  .hdr{{background:#0F2B5B;padding:26px;text-align:center;}}
  .hdr h1{{margin:0;font-size:22px;color:#fff;letter-spacing:3px;font-weight:900;}}
  .hdr p{{margin:6px 0 0;color:#8899bb;font-size:13px;}}
  .gold{{height:3px;background:linear-gradient(90deg,#E8A020,transparent);}}
  .body{{background:#fff;padding:24px;}}
  .sec{{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:1px;color:#0F2B5B;
        border-bottom:2px solid #E8A020;padding-bottom:6px;margin:22px 0 12px;}}
  table{{border-collapse:collapse;}}
  .footer{{background:#071525;padding:14px;text-align:center;}}
  .footer p{{margin:3px 0;font-size:11px;color:#8899bb;}}
  .footer strong{{color:#E8A020;}}
</style></head><body><div class="wrap">
  <div class="hdr"><h1>&#128202; ENERGY COMERCIAL</h1><p>Informe diario &mdash; {date_str}</p></div>
  <div class="gold"></div>
  <div class="body">
    <div class="sec">&#9889; Actividad de hoy</div>
    {hoy_html}
    <div class="sec">&#128200; Mes actual</div>
    <table width='100%' cellpadding='6' cellspacing='6'><tr>
      {_card(k_mes['n'], 'Cotizaciones', '#e8f0fe', '#1A3A8F', '#0F2B5B')}
      {_card(money(k_mes['monto']), 'Cotizado', '#e8f0fe', '#1A3A8F', '#0F2B5B')}
      {_card(f"{k_mes['conv']:.0f}%", 'Conversión', '#e6f4ea', '#2EAA4A', '#1B5E20')}
      {_card(money(k_mes['pipeline']), 'Pipeline abierto', '#fff8e1', '#E8A020', '#b7770d')}
    </tr></table>
    {margen_html}
    <div class="sec">&#128197; Tasa de conversi&oacute;n &mdash; &uacute;ltimos 6 meses</div>
    <table width='100%' cellpadding='0' cellspacing='0' style='font-size:12px;'>
      <thead><tr>
        <th style='background:#0F2B5B;color:#fff;padding:7px 10px;text-align:left;font-size:10px;'>MES</th>
        <th style='background:#0F2B5B;color:#fff;padding:7px 10px;text-align:center;font-size:10px;'>COTIZADAS</th>
        <th style='background:#0F2B5B;color:#fff;padding:7px 10px;text-align:right;font-size:10px;'>$ COTIZADO</th>
        <th style='background:#0F2B5B;color:#fff;padding:7px 10px;text-align:center;font-size:10px;'>GANADAS</th>
        <th style='background:#0F2B5B;color:#fff;padding:7px 10px;text-align:right;font-size:10px;'>$ GANADO</th>
        <th style='background:#0F2B5B;color:#fff;padding:7px 10px;text-align:center;font-size:10px;'>% CONV</th>
      </tr></thead><tbody>{serie_rows}</tbody></table>
    <div class="sec">&#127942; Top clientes del a&ntilde;o (por $ ganado)</div>
    <table width='100%' cellpadding='0' cellspacing='0' style='font-size:12px;'>
      <thead><tr>
        <th style='background:#0F2B5B;color:#fff;padding:7px 10px;text-align:left;font-size:10px;'>CLIENTE</th>
        <th style='background:#0F2B5B;color:#fff;padding:7px 10px;text-align:right;font-size:10px;'>$ GANADO</th>
        <th style='background:#0F2B5B;color:#fff;padding:7px 10px;text-align:center;font-size:10px;'>GANADAS/COTIZADAS</th>
      </tr></thead><tbody>{tops_rows}</tbody></table>
    <div style='background:#f8f9ff;border-radius:8px;padding:10px 14px;font-size:12px;color:#555;margin:16px 0 0;'>
      &#128206; Adjunto va <b>Datos_EYG.xlsx</b> con TODA la informaci&oacute;n
      ({totales['cots']} cotizaciones &middot; {totales['remis']} l&iacute;neas de remisi&oacute;n &middot;
      {totales['planes']} l&iacute;neas de plan de compras &middot; {totales['ocs']} OCs) para tablas din&aacute;micas y an&aacute;lisis.
    </div>
  </div>
  <div class="footer">
    <p>&#9889; Generado por <strong>ENERGY &mdash; Asistente Administrativo</strong></p>
    <p>E&amp;G Energy Group &middot; Informe autom&aacute;tico L&ndash;V 7:00 PM Colombia</p>
  </div>
</div></body></html>"""


# ─── 5. MICROSOFT GRAPH ───────────────────────────────────────────────────────

def get_access_token(tenant_id, client_id, client_secret):
    url = 'https://login.microsoftonline.com/' + tenant_id + '/oauth2/v2.0/token'
    data = urllib.parse.urlencode({
        'grant_type': 'client_credentials', 'client_id': client_id,
        'client_secret': client_secret, 'scope': 'https://graph.microsoft.com/.default'
    }).encode()
    req = urllib.request.Request(url, data=data, method='POST')
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())['access_token']


def send_email(token, sender, recipients, subject, html_body, attachment_bytes=None, attachment_name=None):
    msg = {
        'subject': subject,
        'body': {'contentType': 'HTML', 'content': html_body},
        'toRecipients': [{'emailAddress': {'address': r.strip()}} for r in recipients if r.strip()]
    }
    if attachment_bytes and attachment_name:
        msg['attachments'] = [{
            '@odata.type': '#microsoft.graph.fileAttachment',
            'name': attachment_name,
            'contentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'contentBytes': base64.b64encode(attachment_bytes).decode('utf-8')
        }]
    payload = json.dumps({'message': msg, 'saveToSentItems': True}).encode('utf-8')
    url = 'https://graph.microsoft.com/v1.0/users/' + sender + '/sendMail'
    req = urllib.request.Request(url, data=payload, method='POST',
        headers={'Authorization': 'Bearer ' + token, 'Content-Type': 'application/json'})
    with urllib.request.urlopen(req) as resp:
        print('✅ Correo enviado (HTTP ' + str(resp.status) + ')')


# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    tenant_id     = os.environ.get('MS_TENANT_ID', '').strip()
    client_id     = os.environ.get('MS_CLIENT_ID', '').strip()
    client_secret = os.environ.get('MS_CLIENT_SECRET', '').strip()
    sender_email  = os.environ.get('SENDER_EMAIL', '').strip()
    recipients    = [r.strip() for r in os.environ.get('RECIPIENT_EMAILS',
                     'andrea.bernal@eygenergygroup.com').split(',')]

    print('📥 Cargando datos...')
    hist  = load_json('data/cotizaciones_historicas.json')
    plat  = load_json('data/cotizaciones.json')
    sols  = load_json('data/solicitudes_cotiz.json')
    print(f'   Histórico: {len(hist)} líneas · Plataforma: {len(plat)} cotizaciones · Solicitudes: {len(sols)}')
    remis  = fetch_supabase('remisiones', '&order=remision.desc')
    planes = fetch_supabase('plan_compras', '&order=cc.desc')
    ocs    = fetch_supabase('oc_compras', '&order=oc.desc')
    provs  = fetch_supabase('proveedores', '&order=nombre.asc')

    cots = unify_cotizaciones(hist, plat)
    print(f'📊 Cotizaciones unificadas: {len(cots)}')

    now = now_co()
    ym = now.strftime('%Y-%m')
    k_mes  = kpis_mes(cots, ym)
    serie  = serie_meses(cots, 6)
    tops   = top_clientes(cots, now.year)
    act    = hoy_actividad(plat, remis, planes, ocs, sols)

    # Margen del mes (solo cotizaciones con venta y costo > 0)
    mes_con_costo = [c for c in cots if (c['fecha'] or '')[:7] == ym and c['venta'] > 0 and c['costo'] > 0]
    margen_mes = None
    if mes_con_costo:
        tv = sum(c['venta'] for c in mes_con_costo)
        tc = sum(c['costo'] for c in mes_con_costo)
        if tv > 0:
            margen_mes = (1 - tc / tv) * 100

    day      = DAYS_ES.get(now.strftime('%A'), now.strftime('%A'))
    month    = MONTHS_ES.get(now.strftime('%B'), now.strftime('%B'))
    date_str = f'{day} {now.day} de {month} de {now.year}'

    totales = {'cots': len(cots), 'remis': len(remis), 'planes': len(planes), 'ocs': len(ocs)}
    html_body = generate_html(date_str, act, k_mes, serie, tops, margen_mes, totales)

    excel_bytes = generate_excel(cots, hist, plat, remis, planes, ocs, provs, sols, serie)
    excel_name  = 'Datos_EYG_' + now.strftime('%Y-%m-%d') + '.xlsx'
    if excel_bytes:
        print(f'📊 Excel generado: {excel_name} ({len(excel_bytes)//1024} KB)')

    if not (tenant_id and client_id and client_secret and sender_email):
        # Modo prueba local: guardar archivos en vez de enviar
        out_dir = os.environ.get('TEST_OUT', '.')
        with open(os.path.join(out_dir, 'informe_test.html'), 'w', encoding='utf-8') as f:
            f.write(html_body)
        if excel_bytes:
            with open(os.path.join(out_dir, excel_name), 'wb') as f:
                f.write(excel_bytes)
        print('🧪 Sin credenciales Graph — informe guardado localmente para revisión.')
        sys.exit(0)

    print('🔑 Obteniendo token Microsoft...')
    token = get_access_token(tenant_id, client_id, client_secret)
    subject = '📊 Informe Comercial E&G — ' + date_str
    print('📧 Enviando a: ' + ', '.join(recipients))
    send_email(token, sender_email, recipients, subject, html_body, excel_bytes, excel_name)
    print('🎉 Informe enviado.')
