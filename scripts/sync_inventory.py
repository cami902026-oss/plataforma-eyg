"""
ENERGY вЂ” SincronizaciГіn de Inventario desde OneDrive
=====================================================
Descarga el archivo Excel desde OneDrive vГ­a Microsoft Graph API,
parsea el inventario y actualiza el array INV_RAW en Index.html.

Ejecutado automГЎticamente por GitHub Actions
Lunes a Viernes a las 3:00 PM Colombia (UTC-5 = 20:00 UTC)
"""

import json
import os
import re
import sys
import io
import urllib.request
import urllib.parse
import urllib.error
import openpyxl

# в”Ђв”Ђв”Ђ CONFIGURACIГ“N в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ

# Ruta del archivo Excel en OneDrive (relativa a la raГ­z del drive)
ONEDRIVE_FILE_PATH = 'LOGISTICA/Inventario definitivo_2026.xlsm'

# Email del usuario cuyo OneDrive contiene el archivo
ONEDRIVE_USER_EMAIL = 'andrea.bernal@eygenergygroup.com'

# Mapeo de nombres de columna Excel в†’ nombres internos INV_RAW
# (normalizado a minГєsculas para comparaciГіn flexible)
COLUMN_MAP = {
    'codigo producto':   'CODIGO PRODUCTO',
    'cГіdigo producto':   'CODIGO PRODUCTO',
    'codigo':            'CODIGO PRODUCTO',
    'descripcion':       'DESCRIPCION',
    'descripciГіn':       'DESCRIPCION',
    'marca':             'MARCA',
    'ubicacion':         'UBICACIГ“N',
    'ubicaciГіn':         'UBICACIГ“N',
    'stock actual':      'STOCK ACTUAL',
    'stock':             'STOCK ACTUAL',
    'entradas':          'ENTRADAS',
    'salidas':           'SALIDAS',
    'familia':           'FAMILIA',
    'categoria':         'CATEGORIA',
    'categorГ­a':         'CATEGORIA',
    'lote':              'LOTE',
}


# в”Ђв”Ђв”Ђ 1. AUTENTICACIГ“N MICROSOFT GRAPH в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ

def get_access_token(tenant_id, client_id, client_secret):
    url = ('https://login.microsoftonline.com/' + tenant_id +
           '/oauth2/v2.0/token')
    data = urllib.parse.urlencode({
        'grant_type':    'client_credentials',
        'client_id':     client_id,
        'client_secret': client_secret,
        'scope':         'https://graph.microsoft.com/.default'
    }).encode()
    req = urllib.request.Request(url, data=data, method='POST')
    try:
        with urllib.request.urlopen(req) as resp:
            print('вњ… Token obtenido')
            return json.loads(resp.read())['access_token']
    except urllib.error.HTTPError as e:
        print('ERROR auth: ' + str(e.code) + ' вЂ” ' + e.read().decode())
        sys.exit(1)


# в”Ђв”Ђв”Ђ 2. DIAGNГ“STICO: LISTA CONTENIDO DEL DRIVE в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ

def list_drive_folder(token, user_email, folder_path='root/children'):
    """Lista los archivos/carpetas en el OneDrive del usuario para diagnГіstico."""
    url = 'https://graph.microsoft.com/v1.0/users/' + user_email + '/drive/' + folder_path
    req = urllib.request.Request(url, headers={'Authorization': 'Bearer ' + token})
    try:
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
            items = data.get('value', [])
            print('рџ“Ѓ Contenido de /' + folder_path + ':')
            for item in items:
                tipo = 'рџ“Ѓ' if 'folder' in item else 'рџ“„'
                print('   ' + tipo + ' ' + item.get('name', '?'))
            return items
    except urllib.error.HTTPError as e:
        print('вљ пёЏ  No se pudo listar ' + folder_path + ': ' + str(e.code) + ' ' + e.read().decode())
        return []


# в”Ђв”Ђв”Ђ 3. DESCARGA EXCEL DESDE ONEDRIVE в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ

def download_excel(token, user_email, file_path):
    """Descarga el archivo Excel del OneDrive del usuario viГa Graph API."""
    encoded = urllib.parse.quote(file_path, safe='/')
    url = ('https://graph.microsoft.com/v1.0/users/' + user_email +
           '/drive/root:/' + encoded + ':/content')
    req = urllib.request.Request(
        url, headers={'Authorization': 'Bearer ' + token})
    print('рџ“Ґ Descargando: ' + url)
    try:
        with urllib.request.urlopen(req) as resp:
            data = resp.read()
            print('вњ… Excel descargado (' + str(len(data)) + ' bytes)')
            return data
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print('ERROR descargando Excel: ' + str(e.code) + ' вЂ” ' + body)
        if e.code == 403:
            print('вљ пёЏ  El app necesita permiso Files.Read.All en Azure AD.')
        elif e.code == 404:
            print('вљ пёЏ  Archivo no encontrado. Ruta: ' + file_path)
        sys.exit(1)


# в”Ђв”Ђв”Ђ 3. PARSEA EL EXCEL в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ

def parse_excel(excel_bytes):
    """Lee el Excel y retorna lista de dicts con los productos."""
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active
    print('рџ“‹ Hoja activa: ' + str(ws.title))

    # Buscar la fila de encabezados (puede no estar en la fila 1)
    # Buscamos en las primeras 15 filas la que tenga mГЎs columnas reconocibles
    header_row_idx = 1
    best_score = 0
    for row_idx in range(1, 16):
        row_vals = [str(c.value or '').strip().lower() for c in ws[row_idx]]
        score = sum(1 for v in row_vals if v in COLUMN_MAP)
        non_empty = sum(1 for v in row_vals if v)
        print('   Fila ' + str(row_idx) + ': ' + str([str(c.value or '').strip() for c in ws[row_idx]][:8]) + ' (score=' + str(score) + ')')
        if score > best_score or (score == best_score and non_empty > 2 and score > 0):
            best_score = score
            header_row_idx = row_idx

    print('рџ“Њ Fila de encabezados detectada: ' + str(header_row_idx) + ' (score=' + str(best_score) + ')')
    print('рџ“Њ Hojas disponibles: ' + str(wb.sheetnames) + ' | Hoja activa: ' + str(ws.title))
    raw_headers = [str(c.value or '').strip() for c in ws[header_row_idx]]
    print('рџ“Њ Encabezados RAW del Excel: ' + str(raw_headers))
    headers = []
    for h in raw_headers:
        mapped = COLUMN_MAP.get(h.lower(), h.upper() if h else '')
        headers.append(mapped)

    print('рџ“Њ Columnas mapeadas: ' + str(headers))

    # Verificar columnas mГ­nimas requeridas
    required = {'CODIGO PRODUCTO', 'DESCRIPCION', 'STOCK ACTUAL'}
    found = set(headers)
    missing = required - found
    if missing:
        print('вљ пёЏ  ALERTA вЂ” Columnas no encontradas: ' + str(missing))
        print('   Encabezados del Excel que no coincidieron: ' +
              str([h for h in raw_headers if COLUMN_MAP.get(h.lower()) is None and h]))
    else:
        print('вњ… Todas las columnas requeridas encontradas')

    # Parsear filas (empezar despuГ©s de la fila de encabezados)
    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None or str(v).strip() == '' for v in row):
            continue  # saltar filas vacГ­as
        record = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                record[headers[i]] = str(val).strip() if val is not None else ''
        # Convertir campos numГ©ricos a entero
        for field in ['STOCK ACTUAL', 'ENTRADAS', 'SALIDAS']:
            if field in record:
                try:
                    record[field] = int(float(record[field])) if record[field] else 0
                except (ValueError, TypeError):
                    record[field] = 0
        records.append(record)

    print('вњ… ' + str(len(records)) + ' productos parseados')
    return records


# в”Ђв”Ђв”Ђ 4. ACTUALIZA ARCHIVOS HTML в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ

def update_html_array(records, html_path, array_name):
    """Reemplaza un array JS en un archivo HTML con los datos frescos."""
    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()

    new_json = json.dumps(records, ensure_ascii=False, separators=(',', ':'))
    pattern = r'const ' + array_name + r'\s*=\s*\[[\s\S]*?\];'
    replacement = 'const ' + array_name + ' = ' + new_json + ';'

    if not re.search(pattern, content):
        print('вљ пёЏ  No se encontrГі "const ' + array_name + '" en ' + html_path)
        return False

    new_content = re.sub(pattern, replacement, content)
    if new_content == content:
        print('в„№пёЏ  Sin cambios en ' + array_name)
    else:
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        print('вњ… ' + html_path + ' в†’ ' + array_name + ' actualizado con ' + str(len(records)) + ' productos')
    return True

def update_index_html(records, html_path):
    update_html_array(records, html_path, 'INV_RAW')

def update_buscador_html(records, html_path):
    update_html_array(records, html_path, 'STOCK_DATA')


# в”Ђв”Ђв”Ђ MAIN в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ

if __name__ == '__main__':
    tenant_id     = os.environ.get('MS_TENANT_ID', '').strip()
    client_id     = os.environ.get('MS_CLIENT_ID', '').strip()
    client_secret = os.environ['MS_CLIENT_SECRET'].strip()
    sender_email  = os.environ['SENDER_EMAIL'].strip()

    print('рџ”Ќ Usuario OneDrive: ' + ONEDRIVE_USER_EMAIL)
    print('рџ“‚ Archivo:          ' + ONEDRIVE_FILE_PATH)

    print('рџ”‘ Obteniendo token Microsoft...')
    token = get_access_token(tenant_id, client_id, client_secret)

    # DiagnГіstico: listar raГ­z y carpeta LOGISTICA
    root_items = list_drive_folder(token, ONEDRIVE_USER_EMAIL)
    logistica_items = [i for i in root_items if i.get('name','').upper() == 'LOGISTICA']
    if logistica_items:
        folder_id = logistica_items[0]['id']
        list_drive_folder(token, ONEDRIVE_USER_EMAIL, 'items/' + folder_id + '/children')
    else:
        print('вљ пёЏ  Carpeta LOGISTICA no encontrada en la raГ­z. Buscando variantes...')
        for item in root_items:
            if 'logis' in item.get('name','').lower():
                print('   Encontrado: ' + item.get('name',''))

    excel_bytes = download_excel(token, ONEDRIVE_USER_EMAIL, ONEDRIVE_FILE_PATH)

    records = parse_excel(excel_bytes)
    if not records:
        print('ERROR: El Excel estГЎ vacГ­o o no tiene datos.')
        sys.exit(1)

    base = os.path.dirname(__file__) + '/..'
    update_index_html(records, os.path.join(base, 'Index.html'))
    update_buscador_html(records, os.path.join(base, 'Buscador_Inventario_2026.html'))

    print('рџЋ‰ SincronizaciГіn completada exitosamente.')
